import os
import shutil
import logging
import csv
import json
import stat
import sys
import webbrowser
import copy
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox

from .utils import _import_with_fallback, CaseEncoder
from .exporter import clean_cell_value
from .config import PDFReconConfig
from .chain_of_custody import get_custody_log_path, log_signed_report, sha256_file

openpyxl = _import_with_fallback('openpyxl', 'Workbook', 'openpyxl')
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

class ExportMixin:
    def _write_case_to_file(self, filepath):
        case_data = {
            'app_version': self.app_version,
            'scan_folder': self.last_scan_folder,
            'all_scan_data': self.all_scan_data,
            'file_annotations': self.file_annotations,
            'exif_outputs': self.exif_outputs,
            'timeline_data': self.timeline_data,
            'path_to_id': self.path_to_id,
            'revision_counter': self.revision_counter,
            'evidence_hashes': self.evidence_hashes,
        }
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(case_data, f, cls=CaseEncoder)
            
    def _export_reader(self):
        if not self.all_scan_data or not self.last_scan_folder:
            messagebox.showwarning(self._("case_nothing_to_save_title"), self._("case_nothing_to_save_msg"))
            return
        
        base_path_str = filedialog.askdirectory(title=self._("export_reader_title"))
        if not base_path_str:
            return
        
        failed_file = ""
        dest_folder = Path(base_path_str) / "Export"

        try:
            if dest_folder.exists():
                shutil.rmtree(dest_folder)
            dest_folder.mkdir(exist_ok=True)

            evidence_folder = dest_folder / "Evidence"
            evidence_folder.mkdir(exist_ok=True)

            new_scan_data = {}
            new_exif, new_timeline, new_hashes, path_map = {}, {}, {}, {}

            scan_base_path = self._resolve_case_path(self.last_scan_folder)

            for original_item in self.all_scan_data.values():
                item = copy.deepcopy(original_item)
                original_path_str = str(item['path'])
                original_abs_path = self._resolve_case_path(original_path_str)
                
                if not original_abs_path or not original_abs_path.exists():
                    logging.warning(f"Skipping missing file for export: {original_abs_path}")
                    continue

                try:
                    relative_sub_path = original_abs_path.relative_to(scan_base_path)
                except ValueError:
                    relative_sub_path = Path(original_abs_path.name)
                
                dest_file_path = evidence_folder / relative_sub_path
                dest_file_path.parent.mkdir(parents=True, exist_ok=True)

                failed_file = original_abs_path.name
                shutil.copy2(original_abs_path, dest_file_path)
                os.chmod(dest_file_path, stat.S_IREAD | stat.S_IRGRP | stat.S_IROTH)

                new_relative_path_str = str(Path("Evidence") / relative_sub_path)
                
                new_hashes[new_relative_path_str] = self._hash_file(dest_file_path)
                if original_path_str in self.exif_outputs:
                    new_exif[new_relative_path_str] = self.exif_outputs[original_path_str]
                if original_path_str in self.timeline_data:
                    new_timeline[new_relative_path_str] = self.timeline_data[original_path_str]

                path_map[original_path_str] = new_relative_path_str
                item['path'] = new_relative_path_str
                new_scan_data[new_relative_path_str] = item

            for item in new_scan_data.values():
                if item.get('is_revision'):
                    original_parent_path = str(item.get('original_path'))
                    if original_parent_path in path_map:
                        item['original_path'] = path_map[original_parent_path]

            new_annotations = {}
            for original_path, note in self.file_annotations.items():
                if original_path in path_map:
                    new_annotations[path_map[original_path]] = note

            case_filename = f"case_{datetime.now().strftime('%Y%m%d_%H%M%S')}.prc"
            dest_case_file = dest_folder / case_filename
            failed_file = case_filename
            
            case_payload = {
                'app_version': self.app_version,
                'scan_folder': self.last_scan_folder,
                'all_scan_data': new_scan_data,
                'file_annotations': new_annotations, 
                'exif_outputs': new_exif,
                'timeline_data': new_timeline,
                'path_to_id': self.path_to_id,
                'revision_counter': self.revision_counter,
                'evidence_hashes': new_hashes,
            }
            with open(dest_case_file, 'w', encoding='utf-8') as f:
                json.dump(case_payload, f, cls=CaseEncoder)

            source_exe = Path(sys.executable)
            reader_exe_name = f"{source_exe.stem}_Reader{source_exe.suffix}"
            dest_exe = dest_folder / reader_exe_name
            failed_file = reader_exe_name
            shutil.copy2(source_exe, dest_exe)

            dependencies = ["license.txt", "config.ini", "icon.ico"]
            for dep_name in dependencies:
                source_dep = self._resolve_path(dep_name, base_is_parent=True)
                if source_dep.exists():
                    failed_file = dep_name
                    shutil.copy2(source_dep, dest_folder / dep_name)

            logging.info(f"Reader exported successfully to {dest_folder}")
            if messagebox.askyesno(self._("export_reader_success_title"), self._("export_reader_success_msg")):
                webbrowser.open(dest_folder)

        except Exception as e:
            logging.error(f"Failed to export Reader during operation on '{failed_file}': {e}")
            messagebox.showerror(
                self._("export_reader_error_title"),
                self._("export_reader_error_specific_msg").format(filename=failed_file, e=e)
            )

    def _prompt_and_export(self, file_format):
        if not self.report_data:
            messagebox.showwarning(self._("no_data_to_save_title"), self._("no_data_to_save_message"))
            return
        
        file_types = {
            "xlsx": [("Excel files", "*.xlsx")], "csv": [("CSV files", "*.csv")],
            "json": [("JSON files", "*.json")], "html": [("HTML files", "*.html")]
        }
        file_path = filedialog.asksaveasfilename(defaultextension=f".{file_format}", filetypes=file_types[file_format])
        if not file_path: return

        try:
            export_methods = {
                "xlsx": self._export_to_excel, "csv": self._export_to_csv,
                "json": self._export_to_json, "html": self._export_to_html
            }
            export_methods[file_format](file_path)
            
            if messagebox.askyesno(self._("excel_saved_title"), self._("excel_saved_message")):
                webbrowser.open(os.path.dirname(file_path))

        except Exception as e:
            logging.error(f"Error exporting to {file_format.upper()}: {e}")
            messagebox.showerror(self._("excel_save_error_title"), self._("excel_save_error_message").format(e=e))

    def _export_to_excel(self, file_path):
        import logging
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        logging.info(f"Exporting report to Excel file: {file_path}")

        wb = Workbook()
        ws = wb.active
        ws.title = "PDFRecon Results"

        headers = [self._(key) for key in self.columns_keys]
        if len(headers) >= 10:
            headers[9] = f"{self._('col_indicators')} {self._('excel_indicators_overview')}"

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=clean_cell_value(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        
        ws.freeze_panes = 'A2'

        indicators_by_path = {}
        for item in getattr(self, "all_scan_data", {}).values():
            path_str = str(item.get("path"))
            indicator_dict = item.get("indicator_keys") or {}
            if indicator_dict:
                lines = [self._format_indicator_details(key, details) for key, details in indicator_dict.items()]
                lines = [l for l in lines if l]
                indicators_by_path[path_str] = "• " + "\n• ".join(lines) if lines else ""
            else:
                indicators_by_path[path_str] = ""

        for row_idx, row_data in enumerate(getattr(self, "report_data", []), start=2):
            try:
                path = row_data[4] 
            except IndexError:
                path = ""

            exif_text = self.exif_outputs.get(path, "")
            indicators_full = indicators_by_path.get(path, "")
            note_text = self.file_annotations.get(path, "")

            row_out = list(row_data)
            
            while len(row_out) < len(headers):
                row_out.append("")
            
            row_out[8] = exif_text         
            if indicators_full:
                row_out[9] = indicators_full 
            row_out[10] = note_text        

            for col_idx, value in enumerate(row_out, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=clean_cell_value(value))
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col in ws.columns:
            try:
                max_len = max(len(str(c.value).split('\n')[0]) for c in col if c.value)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
            except (ValueError, TypeError):
                pass

        wb.save(file_path)
        self._sign_export_file(file_path)

    def _sign_export_file(self, file_path: str) -> None:
        """After any export: write .sha256 sidecar, optional .sig, and log to chain of custody."""
        path = Path(file_path)
        if not path.exists():
            return
        report_hash = sha256_file(path)
        sha256_sidecar = path.with_suffix(path.suffix + ".sha256")
        sha256_sidecar.write_text(f"{report_hash}  {path.name}\n", encoding="utf-8")

        signature_info = None
        sign_with_key = getattr(PDFReconConfig, "SIGNING_KEY_PATH", None)
        if sign_with_key:
            from .signed_report import sign_file_detached
            key_path = Path(sign_with_key)
            if key_path.exists():
                try:
                    signature_info = sign_file_detached(path, key_path)
                except Exception as e:
                    logging.warning("Export signing failed: %s", e)

        custody_log = None
        if getattr(self, "case_root_path", None):
            root = Path(self.case_root_path) if self.case_root_path else None
            if root:
                custody_log = get_custody_log_path(root, getattr(self, "current_case_filepath", None))
        if custody_log:
            log_signed_report(
                custody_log,
                path,
                report_hash,
                signature_info=signature_info,
                case_path=getattr(self, "current_case_filepath"),
            )

    def _export_to_csv(self, file_path):
        headers = [self._(key) for key in self.columns_keys]
        
        def _indicators_for_path(path_str: str) -> str:
            rec = self.all_scan_data.get(path_str)
            if not rec: return ""
            indicator_dict = rec.get('indicator_keys') or {}
            if not indicator_dict: return ""

            lines = [self._format_indicator_details(key, details) for key, details in indicator_dict.items()]
            lines = [l for l in lines if l]
            return "; ".join(lines)

        data_for_export = []
        for row_data in self.report_data:
            new_row = list(row_data)
            path = new_row[4] 
            exif_output = self.exif_outputs.get(path, "")
            indicators_full = _indicators_for_path(path)
            note_text = self.file_annotations.get(path, "")
            
            while len(new_row) < len(headers):
                new_row.append("")

            new_row[8] = exif_output      
            if indicators_full:
                new_row[9] = indicators_full 
            new_row[10] = note_text      
  
            data_for_export.append(new_row)

        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data_for_export)
        self._sign_export_file(file_path)

    def _export_to_json(self, file_path):
        scan_data_export = []
        for item in self.all_scan_data.values():
            path_str = str(item['path'])
            item_copy = item.copy()
            item_copy['path'] = path_str 
            if 'original_path' in item_copy:
                item_copy['original_path'] = str(item_copy['original_path'])
            
            if 'indicator_keys' in item_copy:
                serializable_indicators = {}
                for key, details in item_copy['indicator_keys'].items():
                    if 'fonts' in details:
                        serializable_details = details.copy()
                        serializable_details['fonts'] = {k: list(v) for k, v in details['fonts'].items()}
                        serializable_indicators[key] = serializable_details
                    else:
                        serializable_indicators[key] = details
                item_copy['indicator_keys'] = serializable_indicators

            item_copy['exif_data'] = self.exif_outputs.get(path_str, "")
            scan_data_export.append(item_copy)
        
        full_export_payload = {
            'scan_results': scan_data_export,
            'file_annotations': self.file_annotations
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(full_export_payload, f, indent=4, default=str)
        self._sign_export_file(file_path)

    def _export_to_html(self, file_path):
        import html
        html_template = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <title>PDFRecon Report</title>
            <style>
                body {{ font-family: sans-serif; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; word-break: break-all; }}
                th {{ background-color: #f2f2f2; }}
                .red-row {{ background-color: #FFDDDD; }}
                .yellow-row {{ background-color: #FFFFCC; }}
                .blue-row {{ background-color: #CCE5FF; }}
                .purple-row {{ background-color: #E8CCFF; }}
                .gray-row {{ background-color: #E0E0E0; }}
            </style>
        </head>
        <body>
            <h1>PDFRecon Report</h1>
            <p>Generated on {date}</p>
            <table>
                <thead><tr>{headers}</tr></thead>
                <tbody>{rows}</tbody>
            </table>
        </body>
        </html>
        """
        headers = "".join(f"<th>{self._(key)}</th>" for key in self.columns_keys)
        rows = ""
        tag_map = {"red_row": "red-row", "yellow_row": "yellow-row", "blue_row": "blue-row", "purple_row": "purple-row", "gray_row": "gray-row"}
        
        for i, values in enumerate(self.report_data):
            tag_class = ""
            try:
                matching_id = next((item_id for item_id in self.tree.get_children() if self.tree.item(item_id, "values")[4] == values[4]), None)
                if matching_id:
                    tags = self.tree.item(matching_id, "tags")
                    if tags:
                        tag_class = tag_map.get(tags[0], "")
            except (IndexError, StopIteration):
                 pass
            
            path_str = values[4]
            note_text = html.escape(self.file_annotations.get(path_str, "")).replace('\n', '<br>')
            
            row_values = [html.escape(str(v)) for v in values]
            while len(row_values) < len(self.columns_keys):
                row_values.append("")
            row_values[10] = note_text

            rows += f'<tr class="{tag_class}">' + "".join(f"<td>{v}</td>" for v in row_values) + "</tr>"

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_template.format(
                date=datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
                headers=headers,
                rows=rows
            ))
        self._sign_export_file(file_path)