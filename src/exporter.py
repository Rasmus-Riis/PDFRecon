"""
Exporter Module

Handles report generation and file export functionality (Excel, CSV, JSON, HTML).
Extracts all export methods from PDFReconApp for modular architecture.
Phase 5b: Complete exporter extraction with all methods.
"""

import logging
import json
import csv
import html as html_escape_module
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .config import UI_COLORS


def clean_cell_value(value):
    """
    Removes control characters and invalid XML characters from cell values.
    Handles mojibake, BOM characters, and XML control characters.
    
    Args:
        value: Cell value to clean
        
    Returns:
        str: Cleaned cell value
    """
    import re
    if value is None:
        return ""
    s = str(value)
    # Remove illegal XML control characters (allow \t \n \r)
    s = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", s)
    # Remove BOM characters
    if s.startswith("\ufeff") or s.startswith("\ufffe"):
        s = s.lstrip("\ufeff\ufffe")
    # Remove mojibake
    if s.startswith("þÿ") or s.startswith("ÿþ"):
        s = s[2:]
    s = s.replace("\x00", "")
    return s


def format_indicator_details(key: str, details: dict) -> str:
    """
    Formats indicator details as a human-readable string.
    Handles different indicator types and their specific data.
    
    Args:
        key: Indicator name
        details: Indicator details dictionary
        
    Returns:
        str: Formatted indicator string
    """
    if not details:
        return key
    
    if isinstance(details, dict):
        # Handle count-based indicators
        if 'count' in details:
            return f"{key} ({details['count']})"
        # Handle text-based indicators
        if 'text' in details:
            return f"{key}: {details['text'][:50]}..."
        # Handle font indicators
        if 'fonts' in details:
            font_count = len(details['fonts'])
            return f"{key} ({font_count} fonts)"
        # Handle list indicators
        if 'items' in details and isinstance(details['items'], list):
            return f"{key} ({len(details['items'])} items)"
    
    return key


def export_to_excel(file_path, report_data: list, all_scan_data: dict, file_annotations: dict, 
                   exif_outputs: dict, column_keys: list, get_translation=None):
    """
    Exports the displayed data to XLSX with a frozen header and word wrap enabled.
    Includes all indicators, EXIF data, and annotations.
    
    Args:
        file_path: Output file path
        report_data: List of result rows to export
        all_scan_data: Dictionary of all scan data
        file_annotations: Dictionary of file notes
        exif_outputs: Dictionary of EXIF outputs
        column_keys: List of column translation keys
        get_translation: Function to translate column keys (optional)
    """
    try:
        logging.info(f"Exporting report to Excel file: {file_path}")

        wb = Workbook()
        ws = wb.active
        ws.title = "PDFRecon Results"

        # Use translation function if provided, otherwise use raw keys
        if get_translation:
            headers = [get_translation(key) for key in column_keys]
        else:
            headers = column_keys
        
        if len(headers) >= 10:
            headers[9] = f"{headers[9] if get_translation else 'Indicators'} (Overview)"

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=clean_cell_value(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        
        ws.freeze_panes = 'A2'

        # Create a lookup dictionary once to avoid repeated searches (optimization)
        indicators_by_path = {}
        for item in all_scan_data.values():
            path_str = str(item.get("path"))
            indicator_dict = item.get("indicator_keys") or {}
            if indicator_dict:
                lines = [format_indicator_details(key, details) for key, details in indicator_dict.items()]
                indicators_by_path[path_str] = "• " + "\n• ".join(lines)
            else:
                indicators_by_path[path_str] = ""

        for row_idx, row_data in enumerate(report_data, start=2):
            try:
                path = row_data[4]  # Path is at index 4
            except IndexError:
                path = ""

            exif_text = exif_outputs.get(path, "")
            indicators_full = indicators_by_path.get(path, "")
            note_text = file_annotations.get(path, "")

            row_out = list(row_data)
            
            while len(row_out) < len(headers):
                row_out.append("")
            
            row_out[8] = exif_text         # EXIF is at index 8
            if indicators_full:
                row_out[9] = indicators_full # Indicators is at index 9
            row_out[10] = note_text        # Note is at index 10

            for col_idx, value in enumerate(row_out, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=clean_cell_value(value))
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col in ws.columns:
            try:
                max_len = max(len(str(c.value).split('\n')[0]) for c in col if c.value)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
            except (ValueError, TypeError):
                pass

        wb.save(file_path)
        logging.info(f"Excel export completed: {file_path}")
        
    except Exception as e:
        logging.error(f"Error exporting to Excel: {e}")
        raise


def export_to_csv(file_path, report_data: list, all_scan_data: dict, file_annotations: dict,
                 exif_outputs: dict, column_keys: list, get_translation=None):
    """
    Exports the displayed data to a CSV file with EXIF and indicator data.
    
    Args:
        file_path: Output file path
        report_data: List of result rows to export
        all_scan_data: Dictionary of all scan data
        file_annotations: Dictionary of file notes
        exif_outputs: Dictionary of EXIF outputs
        column_keys: List of column translation keys
        get_translation: Function to translate column keys (optional)
    """
    try:
        # Use translation function if provided
        if get_translation:
            headers = [get_translation(key) for key in column_keys]
        else:
            headers = column_keys
        
        def _indicators_for_path(path_str: str) -> str:
            """Helper function to get a semicolon-separated string of indicators."""
            rec = all_scan_data.get(path_str)
            if not rec:
                return ""
            indicator_dict = rec.get('indicator_keys') or {}
            if not indicator_dict:
                return ""
            lines = [format_indicator_details(key, details) for key, details in indicator_dict.items()]
            return "; ".join(lines)

        # Prepare data with full EXIF output + full indicators
        data_for_export = []
        for row_data in report_data:
            new_row = list(row_data)
            path = new_row[4]  # Path is at index 4
            exif_output = exif_outputs.get(path, "")
            indicators_full = _indicators_for_path(path)
            note_text = file_annotations.get(path, "")
            
            while len(new_row) < len(headers):
                new_row.append("")

            new_row[8] = exif_output      # EXIF is at index 8
            if indicators_full:
                new_row[9] = indicators_full # Indicators is at index 9
            new_row[10] = note_text       # Note is at index 10
  
            data_for_export.append(new_row)

        # Use utf-8-sig for better Excel compatibility with special characters
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data_for_export)
        
        logging.info(f"CSV export completed: {file_path}")
        
    except Exception as e:
        logging.error(f"Error exporting to CSV: {e}")
        raise


def export_to_json(file_path, all_scan_data: dict, file_annotations: dict, exif_outputs: dict):
    """
    Exports a more detailed report of all scanned data and notes to a JSON file.
    Includes indicator details, EXIF data, and annotations.
    
    Args:
        file_path: Output file path
        all_scan_data: Dictionary of all scan data
        file_annotations: Dictionary of file notes
        exif_outputs: Dictionary of EXIF outputs
    """
    try:
        scan_data_export = []
        for item in all_scan_data.values():
            path_str = str(item['path'])
            item_copy = item.copy()
            item_copy['path'] = path_str  # Convert Path object to string
            if 'original_path' in item_copy:
                item_copy['original_path'] = str(item_copy['original_path'])
            
            if 'indicator_keys' in item_copy:
                serializable_indicators = {}
                for key, details in item_copy['indicator_keys'].items():
                    if 'fonts' in details:
                        serializable_details = details.copy()
                        serializable_details['fonts'] = {k: list(v) for k, v in details['fonts'].items()}
                        serializable_indicators[key] = serializable_details
                    else:
                        serializable_indicators[key] = details
                item_copy['indicator_keys'] = serializable_indicators

            item_copy['exif_data'] = exif_outputs.get(path_str, "")
            scan_data_export.append(item_copy)
        
        full_export_payload = {
            'scan_results': scan_data_export,
            'file_annotations': file_annotations
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(full_export_payload, f, indent=4, default=str)
        
        logging.info(f"JSON export completed: {file_path}")
            
    except Exception as e:
        logging.error(f"Error exporting to JSON: {e}")
        raise


def export_to_html(file_path, report_data: list, file_annotations: dict, all_scan_data: dict,
                  column_keys: list, tree_get_children=None, tree_item=None, tag_map=None, get_translation=None):
    """
    Exports a simple, color-coded HTML report with indicators and notes.
    
    Args:
        file_path: Output file path
        report_data: List of result rows to export
        file_annotations: Dictionary of file notes
        all_scan_data: Dictionary of all scan data
        column_keys: List of column translation keys
        tree_get_children: Function to get tree children (optional)
        tree_item: Function to get tree item values (optional)
        tag_map: Dictionary mapping tag names to CSS classes (optional)
        get_translation: Function to translate column keys (optional)
    """
    try:
        # Use translation function if provided
        if get_translation:
            headers_list = [get_translation(key) for key in column_keys]
        else:
            headers_list = column_keys
        
        headers = "".join(f"<th>{h}</th>" for h in headers_list)
        
        if not tag_map:
            tag_map = {"red_row": "red-row", "yellow_row": "yellow-row", "blue_row": "blue-row", "gray_row": "gray-row"}
        
        rows = ""
        
        # Generate Table Rows
        for i, values in enumerate(report_data):
            tag_class = ""
            try:
                # Try to get tag from tree if functions provided
                if tree_get_children and tree_item:
                    matching_id = next((item_id for item_id in tree_get_children() 
                                      if tree_item(item_id, "values")[4] == values[4]), None)
                    if matching_id:
                        tags = tree_item(matching_id, "tags")
                        if tags:
                            tag_class = tag_map.get(tags[0], "")
            except (IndexError, StopIteration, TypeError):
                pass
            
            path_str = values[4]
            note_text = html_escape_module.escape(file_annotations.get(path_str, "")).replace('\n', '<br>')
            
            row_values = [html_escape_module.escape(str(v)) for v in values]
            while len(row_values) < len(headers_list):
                row_values.append("")
            if len(row_values) > 10:
                row_values[10] = note_text

            rows += f'<tr class="{tag_class}">' + "".join(f"<td>{v}</td>" for v in row_values) + "</tr>"

        html_template = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>PDFRecon Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        table {{ border-collapse: collapse; width: 100%; border: 1px solid #ddd; }}
        th {{ background-color: #f2f2f2; padding: 12px; text-align: left; font-weight: bold; border: 1px solid #ddd; }}
        td {{ padding: 8px; border: 1px solid #ddd; word-break: break-word; }}
        .red-row {{ background-color: #FFDDDD; }}
        .yellow-row {{ background-color: #FFFFCC; }}
        .blue-row {{ background-color: #CCE5FF; }}
        .gray-row {{ background-color: #E0E0E0; }}
        h1 {{ color: #333; }}
        .report-date {{ color: #666; font-style: italic; }}
    </style>
</head>
<body>
    <h1>PDFRecon Report</h1>
    <p class="report-date">Generated on {date}</p>
    <table>
        <thead><tr>{headers}</tr></thead>
        <tbody>{rows}</tbody>
    </table>
</body>
</html>
"""
        
        html_content = html_template.format(
            date=datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
            headers=headers,
            rows=rows
        )
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        logging.info(f"HTML export completed: {file_path}")
        
    except Exception as e:
        logging.error(f"Error exporting to HTML: {e}")
        raise
