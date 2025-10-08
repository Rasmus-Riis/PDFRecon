import tkinter as tk
from tkinter import ttk, filedialog, messagebox, Toplevel
import hashlib
import os
import re
import shutil
import subprocess
import zlib
from pathlib import Path
from datetime import datetime, timezone
import threading
import queue
import difflib
import stat
import base64
import copy
import binascii
import webbrowser
import sys
import logging
import tempfile
import multiprocessing
import pickle
from concurrent.futures import ThreadPoolExecutor, as_completed
import configparser
import csv
import json
import time

# --- Optional library imports with error handling ---
try:
    from PIL import Image, ImageTk, ImageChops, ImageOps
except ImportError:
    messagebox.showerror("Missing Library", "The Pillow library is not installed.\n\nPlease run 'pip install Pillow' in your terminal to use this program.")
    sys.exit(1)

try:
    import fitz  # PyMuPDF
except ImportError:
    messagebox.showerror("Missing Library", "PyMuPDF is not installed.\n\nPlease run 'pip install PyMuPDF' in your terminal to use this program.")
    sys.exit(1)

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except ImportError:
    messagebox.showerror("Missing Library", "tkinterdnd2 is not installed.\n\nPlease run 'pip install tkinterdnd2' in your terminal to use this program.")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    messagebox.showerror("Missing Library", "The openpyxl library is not installed.\n\nPlease run 'pip install openpyxl' in your terminal to use this program.")
    sys.exit(1)

try:
    import requests
except ImportError:
    messagebox.showerror("Missing Library", "The requests library is not installed.\n\nPlease run 'pip install requests' in your terminal to use the update checker.")
    sys.exit(1)
    
# --- OCG (layers) detection helpers ---
_LAYER_OCGS_BLOCK_RE = re.compile(rb"/OCGs\s*\[(.*?)\]", re.S)
_OBJ_REF_RE          = re.compile(rb"(\d+)\s+(\d+)\s+R")
_LAYER_OC_REF_RE     = re.compile(rb"/OC\s+(\d+)\s+(\d+)\s+R")

def count_layers(pdf_bytes: bytes) -> int:
    """
    Conservatively counts OCGs (layers) in PDF bytes.
    1) Finds /OCGs [ ... ] and collects all indirect refs "n m R".
    2) Also finds /OC n m R in content/resources.
    3) Deduplicates (n, gen).
    """
    refs = set()

    m = _LAYER_OCGS_BLOCK_RE.search(pdf_bytes)
    if m:
        for n, g in _OBJ_REF_RE.findall(m.group(1)):
            refs.add((int(n), int(g)))

    for n, g in _LAYER_OC_REF_RE.findall(pdf_bytes):
        refs.add((int(n), int(g)))

    return len(refs)


# --- PHASE 1/3: Configuration and Custom Exceptions ---
class PDFReconConfig:
    """Configuration settings for PDFRecon. Values are loaded from config.ini."""
    MAX_FILE_SIZE = 1000 * 1024 * 1024  # 500MB
    MAX_REVISIONS = 100
    EXIFTOOL_TIMEOUT = 30
    MAX_WORKER_THREADS = min(16, (os.cpu_count() or 4) * 2)
    VISUAL_DIFF_PAGE_LIMIT = 15
    EXPORT_INVALID_XREF = False

class PDFProcessingError(Exception):
    """Base exception for PDF processing errors."""
    pass

class PDFCorruptionError(PDFProcessingError):
    """Exception for corrupt or unreadable PDF files."""
    pass

class PDFTooLargeError(PDFProcessingError):
    """Exception for files that exceed the size limit."""
    pass

class PDFEncryptedError(PDFProcessingError):
    """Exception for encrypted files that cannot be read."""
    pass
# --- End Phase 1/3 ---
def md5_file(fp: Path, buf_size: int = 4 * 1024 * 1024) -> str:
    """
    Fast MD5 with reusable buffer (fewer allocations).
    """
    h = hashlib.md5()
    with fp.open("rb", buffering=0) as f:
        buf = bytearray(buf_size)
        mv = memoryview(buf)
        while True:
            n = f.readinto(mv)
            if not n:
                break
            h.update(mv[:n])
    return h.hexdigest()


def fmt_times_pair(ts: float):
    """Return ('DD-MM-YYYY HH:MM:SS±ZZZZ', 'YYYY-mm-ddTHH:MM:SSZ')."""
    local = datetime.fromtimestamp(ts).astimezone()
    utc = datetime.fromtimestamp(ts, tz=timezone.utc)
    return local.strftime("%d-%m-%Y %H:%M:%S%z"), utc.strftime("%Y-%m-%dT%H:%M:%SZ")

def safe_stat_times(path: Path):
    try:
        st = path.stat()
        return st.st_ctime, st.st_mtime
    except Exception:
        return None, None



class PDFReconApp:

    def __init__(self, root):
        # --- Application Configuration ---
        self.app_version = "16.9.0"
        self.config_path = self._resolve_path("config.ini", base_is_parent=True)
        self._load_or_create_config()
        
        # --- Root Window Setup ---
        self.root = root
        
        self.is_reader_mode = "reader" in Path(sys.executable).stem.lower()
        title = f"PDFRecon Reader v{self.app_version}" if self.is_reader_mode else f"PDFRecon v{self.app_version}"
        self.base_title = title
        self.root.title(title)
        self.root.geometry("1200x700")
        self.inspector_window = None
        self.inspector_doc = None
        self.inspector_pdf_update_job = None

        # --- Set Application Icon ---
        try:
            icon_path = self._resolve_path('icon.ico')
            if icon_path.exists():
                self.root.iconbitmap(icon_path)
            else:
                logging.warning("icon.ico not found. Using default icon.")
        except tk.TclError:
            logging.warning("Could not load icon.ico. Using default icon.")
        except Exception as e:
            logging.error(f"Unexpected error when loading icon: {e}")

        # --- Application Data ---
        self.report_data = [] 
        self.all_scan_data = []
        self.last_scan_folder = None 
        self.case_root_path = None
        self.current_case_filepath = None
        self.file_annotations = {}
        self.dirty_notes = set()
        self.evidence_hashes = {}
        self.exif_outputs = {}
        self.timeline_data = {}
        self.path_to_id = {}
        self.scan_start_time = 0

        # --- State Variables ---
        self.revision_counter = 0
        self.scan_queue = queue.Queue()
        self.copy_executor = None
        self.case_is_dirty = False       
        self.tree_sort_column = None
        self.tree_sort_reverse = False
        self.exif_popup = None
        self.indicator_popup = None

        # --- Language and Filter Setup ---
        self.language = tk.StringVar(value=self.default_language)
        self.filter_var = tk.StringVar()
        
        # --- Compile regex for software detection once ---
        self.software_tokens = re.compile(
            r"(adobe|acrobat|billy|businesscentral|cairo|canva|chrome|chromium|clibpdf|dinero|dynamics|economic|edge|eboks|excel|firefox|"
            r"formpipe|foxit|fpdf|framemaker|ghostscript|illustrator|indesign|ilovepdf|itext|"
            r"kmd|lasernet|latex|libreoffice|microsoft|navision|netcompany|nitro|office|openoffice|pdflatex|pdf24|photoshop|powerpoint|prince|"
            r"quartz|reportlab|safari|skia|tcpdf|tex|visma|word|wkhtml|wkhtmltopdf|xetex)",
            re.IGNORECASE
        )
        
        # --- GUI Setup ---
        self._setup_logging()
        self.translations = self.get_translations() 
        self._setup_styles()
        self._setup_menu()
        self._setup_main_frame()
        self._setup_drag_and_drop()
        
        logging.info(f"PDFRecon v{self.app_version} started in {'Reader' if self.is_reader_mode else 'Full'} mode.")

        # --- NY LOGIK: Auto-load case in Reader mode ---
        if self.is_reader_mode:
            self.root.after(100, self._autoload_case_in_reader)
            
    def _(self, key):
        """Returns the translated text for a given key."""
        # Fallback for keys that might not exist in a language
        return self.translations[self.language.get()].get(key, key)

    def get_translations(self):
        """Loads all translations for the application from external files."""
        # Define paths to the language files relative to the script
        base_path = Path(__file__).parent
        json_path = base_path / "lang" / "translations.json"
        manual_da_path = base_path / "lang" / "manual_da.md"
        manual_en_path = base_path / "lang" / "manual_en.md"

        translations = {}

        # Load the main UI translations from the JSON file
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                translations = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError) as e:
            logging.error(f"Could not load or parse translations.json: {e}")
            # Create a fallback structure to prevent crashes
            translations = {"da": {}, "en": {}}

        # Load the Danish manual and inject it into the translations dictionary
        try:
            with open(manual_da_path, 'r', encoding='utf-8') as f:
                translations.get("da", {})["full_manual"] = f.read()
        except FileNotFoundError:
            logging.error(f"Danish manual not found at {manual_da_path}")
            translations.get("da", {})["full_manual"] = "Manual not found."

        # Load the English manual and inject it into the translations dictionary
        try:
            with open(manual_en_path, 'r', encoding='utf-8') as f:
                translations.get("en", {})["full_manual"] = f.read()
        except FileNotFoundError:
            logging.error(f"English manual not found at {manual_en_path}")
            translations.get("en", {})["full_manual"] = "Manual not found."
            
        # Ensure 'about_version' is dynamically inserted
        version_string = f"PDFRecon v{self.app_version}"
        if "da" in translations:
            translations["da"]["about_version"] = version_string
        if "en" in translations:
            translations["en"]["about_version"] = version_string

        return translations
  
    def _load_or_create_config(self):
        """Loads configuration from config.ini or creates the file with default values."""
        parser = configparser.ConfigParser()
        # Set a default language in case the config file is missing or corrupt.
        self.default_language = "en"
        if not self.config_path.exists():
            logging.info("config.ini not found. Creating with default values.")
            parser['Settings'] = {
                'MaxFileSizeMB': '500',
                'ExifToolTimeout': '30',
                'MaxWorkerThreads': str(PDFReconConfig.MAX_WORKER_THREADS),
                'Language': self.default_language,
                 'VisualDiffPageLimit': str(PDFReconConfig.VISUAL_DIFF_PAGE_LIMIT),
                'ExportInvalidXREF': 'False'
            }
            try:
                with open(self.config_path, 'w') as configfile:
                    configfile.write("# PDFRecon Configuration File\n")
                    parser.write(configfile)
            except IOError as e:
                logging.error(f"Could not write to config.ini: {e}")
                return
        
        try:
            parser.read(self.config_path)
            settings = parser['Settings']
            PDFReconConfig.MAX_FILE_SIZE = settings.getint('MaxFileSizeMB', 500) * 1024 * 1024
            PDFReconConfig.EXIFTOOL_TIMEOUT = settings.getint('ExifToolTimeout', 30)
            PDFReconConfig.MAX_WORKER_THREADS = settings.getint('MaxWorkerThreads', PDFReconConfig.MAX_WORKER_THREADS)
            PDFReconConfig.VISUAL_DIFF_PAGE_LIMIT = settings.getint('VisualDiffPageLimit', 5)
            PDFReconConfig.EXPORT_INVALID_XREF = settings.getboolean('ExportInvalidXREF', False)
            # Load the language, fallback to the default 'en'
            self.default_language = settings.get('Language', 'en')
            logging.info(f"Configuration loaded from {self.config_path}")
        except Exception as e:
            logging.error(f"Could not read config.ini, using default values. Error: {e}")

    def _setup_logging(self):
        """ Sets up a robust logger that writes to a file. """
        self.log_file_path = self._resolve_path("pdfrecon.log", base_is_parent=True)
        
        logger = logging.getLogger()
        logger.setLevel(logging.INFO)
        
        # Clear existing handlers to avoid duplicate logs
        if logger.hasHandlers():
            logger.handlers.clear()
            
        try:
            # Create a file handler to write logs to a file
            fh = logging.FileHandler(self.log_file_path, mode='a', encoding='utf-8')
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            fh.setFormatter(formatter)
            logger.addHandler(fh)
        except Exception as e:
            messagebox.showerror("Log Error", f"Could not create the log file.\n\nDetails: {e}")
    def _autoload_case_in_reader(self):
        """If in reader mode, finds and opens a .prc file in the executable's directory."""
        try:
            exe_dir = Path(sys.executable).parent
            case_files = list(exe_dir.glob('*.prc'))
            
            # Auto-load only if exactly one case file is found
            if len(case_files) == 1:
                logging.info(f"Reader mode: Found case file to auto-load: {case_files[0]}")
                self._open_case(filepath=case_files[0])
            elif len(case_files) > 1:
                logging.warning("Reader mode: Found multiple .prc files. Aborting auto-load.")
            else:
                logging.info("Reader mode: No .prc file found for auto-loading.")
        except Exception as e:
            logging.error(f"Error during case auto-load: {e}")
            
    def _finalize_copy_operations(self):
        """Waits for the copy executor to finish and updates the status bar."""
        if self.copy_executor:
            self.copy_executor.shutdown(wait=True)
            self.copy_executor = None
            logging.info("All background copy operations have finished.")
            # Schedule the final status update on the main GUI thread
            self.root.after(0, self._update_summary_status)

    def _setup_styles(self):
        """Initializes and configures the styles for ttk widgets."""
        self.style = ttk.Style()
        # Use your theme as before (adjust if you had another one)
        try:
            self.style.theme_use("clam")
        except Exception:
            pass

        # Highlight on selection (keep as you had it)
        self.style.map('Treeview', background=[('selected', '#0078D7')])

        # Define the actual colors for tags
        self.style.configure("red_row", background="#FFD6D6")     # reddish background
        self.style.configure("yellow_row", background="#FFF4CC")  # yellow background

        # Map only these statuses now
        self.tree_tags = {
            "JA": "red_row",
            "YES": "red_row",
            "Sandsynligt": "yellow_row",
            "Possible": "yellow_row",
        }
        
    def _update_title(self):
        """Updates the main window title to reflect the current state (case name, unsaved changes)."""
        title = self.base_title
        if self.current_case_filepath:
            title += f" - [{Path(self.current_case_filepath).name}]"
        
        if self.case_is_dirty:
            title += " *"
        
        self.root.title(title)    
    
    def _resolve_case_path(self, path_from_case):
        """Resolves a path from a case file, handling both absolute and relative paths."""
        if not path_from_case:
            return None
        p = Path(path_from_case)
        if p.is_absolute():
            return p
        # If a case is loaded, resolve the path relative to the case's folder.
        if self.case_root_path:
            return self.case_root_path / p
        # Fallback if no case root is set (should not happen in practice)
        return p    

    def _setup_menu(self):
        """Creates the main menu bar for the application."""
        self.menubar = tk.Menu(self.root)
        
        # --- File Menu ---
        self.file_menu = tk.Menu(self.menubar, tearoff=0)
        self.menubar.add_cascade(label=self._("menu_file"), menu=self.file_menu)
        self.file_menu.add_command(label=self._("menu_open_case"), command=self._open_case)
        self.file_menu.add_command(label=self._("menu_verify_integrity"), command=self._verify_integrity, state="disabled")
        
        # Add save/export options only if not in reader mode
        if self.is_reader_mode:
            self.file_menu.add_command(label=self._("menu_save_case_simple"), command=self._save_current_case, state="disabled")
        elif not self.is_reader_mode:
            self.file_menu.add_command(label=self._("menu_save_case"), command=self._save_case, state="disabled")
            
            # The 'Export Reader' menu item is only created if the application is a frozen executable.
            if getattr(sys, 'frozen', False):
                self.file_menu.add_command(label=self._("menu_export_reader"), command=self._export_reader, state="disabled")
            
            self.file_menu.add_separator()
            self.file_menu.add_command(label=self._("menu_settings"), command=self.open_settings_popup)
        
        self.file_menu.add_separator()
        self.file_menu.add_command(label=self._("menu_exit"), command=self.root.quit)

        # --- Help Menu (unchanged) ---
        self.help_menu = tk.Menu(self.menubar, tearoff=0)
        self.lang_menu = tk.Menu(self.help_menu, tearoff=0) 

        self.menubar.add_cascade(label=self._("menu_help"), menu=self.help_menu)
        self.help_menu.add_command(label=self._("menu_manual"), command=self.show_manual)
        self.help_menu.add_command(label=self._("menu_about"), command=self.show_about)
        self.help_menu.add_separator()
        # Add the update check command here
        self.help_menu.add_command(label=self._("menu_check_for_updates"), command=self._check_for_updates)
        self.help_menu.add_separator()
        self.help_menu.add_cascade(label=self._("menu_language"), menu=self.lang_menu)
        self.lang_menu.add_radiobutton(label="Dansk", variable=self.language, value="da", command=self.switch_language)
        self.lang_menu.add_radiobutton(label="English", variable=self.language, value="en", command=self.switch_language)
        self.help_menu.add_separator()
        self.help_menu.add_command(label=self._("menu_license"), command=self.show_license)
        self.help_menu.add_command(label=self._("menu_log"), command=self.show_log_file)
        
        self.root.config(menu=self.menubar)
       
    def _update_summary_status(self):
        """Updates the status bar with a summary of the results from the full scan."""
        if not self.all_scan_data:
            self.status_var.set(self._("status_initial"))
            return

        # Build a temporary list of flags for all scanned files
        all_flags = []
        for data in self.all_scan_data:
            if data.get("status") == "error":
                error_type_key = data.get("error_type", "unknown_error")
                all_flags.append(self._(error_type_key))
            elif not data.get("is_revision"):
                flag = self.get_flag(data.get("indicator_keys", {}), False)
                all_flags.append(flag)

        # Define the set of error statuses for counting
        error_keys = ["file_too_large", "file_corrupt", "file_encrypted", "validation_error", "processing_error", "unknown_error"]
        error_statuses = {self._(key) for key in error_keys}
        
        # Count occurrences of each status type
        changed_count = all_flags.count("JA") + all_flags.count("YES")
        indications_found_count = all_flags.count("Sandsynligt") + all_flags.count("Possible")
        total_altered = changed_count + indications_found_count
                           
        error_count = sum(1 for flag in all_flags if flag in error_statuses)
        
        original_files_count = len([d for d in self.all_scan_data if not d.get('is_revision')])
        # Correctly calculate clean files
        not_flagged_count = original_files_count - changed_count - indications_found_count - error_count

        # Format the summary text based on whether errors were found
        if error_count > 0:
            summary_text = self._("scan_complete_summary_with_errors").format(
                total=original_files_count, total_altered=total_altered,
                changed_count=changed_count, revs=self.revision_counter,
                indications_found_count=indications_found_count, errors=error_count, clean=not_flagged_count
            )
        else:
            summary_text = self._("scan_complete_summary").format(
                total=original_files_count, total_altered=total_altered,
                changed_count=changed_count, revs=self.revision_counter,
                indications_found_count=indications_found_count, clean=not_flagged_count
            )
        self.status_var.set(summary_text)

    def _perform_copy(self, source, dest_path):
        """A simple worker task to copy a file, designed to be run in a thread pool."""
        try:
            dest_path.parent.mkdir(parents=True, exist_ok=True)
            if isinstance(source, Path):
                shutil.copy2(source, dest_path)
                logging.info(f"Copied file from {source.name} to: {dest_path}")
            elif isinstance(source, bytes):
                dest_path.write_bytes(source)
                logging.info(f"Copied revision bytes to: {dest_path}")
        except Exception as e:
            logging.error(f"Error copying to {dest_path}: {e}")
            
    def switch_language(self):
        """Updates all text in the GUI to the selected language."""
        # --- Preserve Selection ---
        path_of_selected = None
        if self.tree.selection():
            selected_item_id = self.tree.selection()[0]
            try:
                path_of_selected = self.tree.item(selected_item_id, "values")[4]
            except IndexError:
                path_of_selected = None # Handle cases where selection is invalid or columns changed

        # --- Re-create the entire menu to ensure all labels are correct ---
        if self.menubar:
            self.menubar.destroy()
        self._setup_menu()

        # --- Update Other GUI Elements ---
        self.scan_button.config(text=self._("choose_folder"))
        self.export_menubutton.config(text=self._("export_report"))
        self.filter_label.config(text=self._("filter_label"))
        
        # --- Update Treeview Column Headers ---
        for i, key in enumerate(self.columns_keys):
            self.tree.heading(self.columns[i], text=self._(key))

        # Re-apply the filter to update the table contents with the new language.
        self._apply_filter() 

        # --- Restore Selection and Details ---
        if path_of_selected:
            new_item_to_select = next((item_id for item_id in self.tree.get_children("") if self.tree.item(item_id, "values")[4] == path_of_selected), None)
            if new_item_to_select:
                self.tree.selection_set(new_item_to_select)
                self.tree.focus(new_item_to_select)
                self.on_select_item(None)
        else:
            self.detail_text.config(state="normal")
            self.detail_text.delete("1.0", tk.END)
            self.detail_text.config(state="disabled")

        # --- Update Status Bar ---
        is_scan_finished = self.scan_button['state'] == 'normal'
        if is_scan_finished and self.all_scan_data:
            self._update_summary_status()
        elif not self.all_scan_data:
            self.status_var.set(self._("status_initial"))

        # --- Re-apply correct menu item states after rebuilding ---
        if self.all_scan_data:
            if self.evidence_hashes:
                self.file_menu.entryconfig(1, state="normal") # Verify File Integrity
            
            if self.is_reader_mode:
                if self.case_is_dirty:
                    self.file_menu.entryconfig(2, state="normal") # Save Case
            else: # Full mode
                self.file_menu.entryconfig(2, state="normal") # Save Case As...
                if getattr(sys, 'frozen', False):
                    self.file_menu.entryconfig(3, state="normal") # Export Reader...

        # --- Save the selected language to the config file ---
        try:
            parser = configparser.ConfigParser()
            parser.read(self.config_path)
            if 'Settings' not in parser:
                parser['Settings'] = {}
            parser['Settings']['Language'] = self.language.get()
            with open(self.config_path, 'w') as configfile:
                configfile.write("# PDFRecon Configuration File\n")
                parser.write(configfile)
            logging.info(f"Language setting saved to {self.config_path}")
        except Exception as e:
            logging.error(f"Could not save language setting to config.ini: {e}")
            
    def _setup_main_frame(self):
        """Sets up the main user interface components within the root window."""
        # --- Main Container Frame ---
        frame = ttk.Frame(self.root, padding=10)
        frame.pack(fill="both", expand=True)

        # --- Top Controls Frame (Scan Button, Filter) ---
        top_controls_frame = ttk.Frame(frame)
        top_controls_frame.pack(pady=5, fill="x")

        self.scan_button = ttk.Button(top_controls_frame, text=self._("choose_folder"), width=25, command=self.choose_folder)
        self.scan_button.pack(side="left", padx=(0, 10))
        if self.is_reader_mode:
            self.scan_button.config(state="disabled")

        self.filter_label = ttk.Label(top_controls_frame, text=self._("filter_label"))
        self.filter_label.pack(side="left", padx=(5, 2))
        
        filter_entry = ttk.Entry(top_controls_frame, textvariable=self.filter_var)
        filter_entry.pack(side="left", fill="x", expand=True)
        self.filter_var.trace_add("write", self._apply_filter)

        # --- Status Bar ---
        initial_status = self._("status_initial_reader") if self.is_reader_mode else self._("status_initial")
        self.status_var = tk.StringVar(value=initial_status)
        status_label = ttk.Label(frame, textvariable=self.status_var, foreground="darkgreen", anchor="w")
        status_label.pack(pady=(5, 10), fill="x", expand=True)

        # --- Treeview (Main Results Table) ---
        tree_frame = ttk.Frame(frame)
        tree_frame.pack(fill="both", expand=True)

        self.columns = ["ID", "Name", "Altered", "Revisions", "Path", "MD5", "File Created", "File Modified", "EXIFTool", "Signs of Alteration", "Note"]
        self.columns_keys = ["col_id", "col_name", "col_changed", "col_revisions", "col_path", "col_md5", "col_created", "col_modified", "col_exif", "col_indicators", "col_note"]
        self.tree = ttk.Treeview(tree_frame, columns=self.columns, show="headings", selectmode="browse")
        
        # Configure row colors
        self.tree.tag_configure("red_row", background='#FFDDDD')
        self.tree.tag_configure("yellow_row", background='#FFFFCC')
        self.tree.tag_configure("blue_row", background='#CCE5FF')
        self.tree.tag_configure("gray_row", background='#E0E0E0')
        
        # Set up treeview columns and headings
        for i, key in enumerate(self.columns_keys):
            self.tree.heading(self.columns[i], text=self._(key), command=lambda c=self.columns[i]: self._sort_column(c, False))
            self.tree.column(self.columns[i], anchor="w", width=120)
        
        self.tree.column("ID", width=40, anchor="center")
        self.tree.column("Name", width=150)
        self.tree.column("Altered", width=100, anchor="w")
        self.tree.column("Revisions", width=80, anchor="center")
        self.tree.column("Note", width=50, anchor="center")

        # Add a scrollbar to the treeview
        tree_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scrollbar.set)
        
        tree_scrollbar.pack(side="right", fill="y")
        self.tree.pack(side="left", fill="both", expand=True)
        
        # Bind events
        self.tree.bind("<<TreeviewSelect>>", self.on_select_item)
        self.tree.bind("<Double-1>", self.show_inspector_popup)
        self.tree.bind("<Motion>", self._on_tree_motion)
        self.tree.bind("<Button-3>", self.show_context_menu)

        # --- Scrollable Details Box ---
        detail_frame = ttk.Frame(frame)
        detail_frame.pack(fill="both", expand=False, pady=(10, 5))

        detail_scrollbar = ttk.Scrollbar(detail_frame, orient="vertical")
        self.detail_text = tk.Text(detail_frame, height=10, wrap="word", font=("Segoe UI", 9),
                                   yscrollcommand=detail_scrollbar.set)
        detail_scrollbar.config(command=self.detail_text.yview)

        self.detail_text.pack(side="left", fill="both", expand=True)
        detail_scrollbar.pack(side="right", fill="y")

        # Configure tags for text formatting in the details box
        self.detail_text.tag_configure("bold", font=("Segoe UI", 9, "bold"))
        self.detail_text.tag_configure("link", foreground="blue", underline=True)
        self.detail_text.tag_bind("link", "<Enter>", lambda e: self.detail_text.config(cursor="hand2"))
        self.detail_text.tag_bind("link", "<Leave>", lambda e: self.detail_text.config(cursor=""))
        self.detail_text.tag_bind("link", "<Button-1>", self._open_path_from_detail)
        
        # --- Bottom Frame (Export Button and Progress Bar) ---
        bottom_frame = ttk.Frame(frame)
        bottom_frame.pack(fill="x", pady=(5,0))

        self.export_menubutton = ttk.Menubutton(bottom_frame, text=self._("export_report"), width=25)
        self.export_menubutton.pack(side="right", padx=5)
        self.export_menu = tk.Menu(self.export_menubutton, tearoff=0)
        self.export_menubutton["menu"] = self.export_menu
        self.export_menu.add_command(label="Excel (.xlsx)", command=lambda: self._prompt_and_export("xlsx"))
        self.export_menu.add_command(label="CSV (.csv)", command=lambda: self._prompt_and_export("csv"))
        self.export_menu.add_command(label="JSON (.json)", command=lambda: self._prompt_and_export("json"))
        self.export_menu.add_command(label="HTML (.html)", command=lambda: self._prompt_and_export("html"))
        self.export_menubutton.config(state="disabled")

        self.progressbar = ttk.Progressbar(bottom_frame, orient="horizontal", mode="determinate", style="blue.Horizontal.TProgressbar")
        self.progressbar.pack(side="left", fill="x", expand=True, padx=5)
        
       
    def _show_note_popup(self):
        """Opens a popup to add or edit a note for the selected file."""
        selected_items = self.tree.selection()
        if not selected_items:
            return
        
        item_id = selected_items[0]
        # This function gets the unique path for the selected row at the moment you right-click.
        # It is critical that column index 4 is the 'Path'.
        path_str = self.tree.item(item_id, "values")[4]
        file_name = self.tree.item(item_id, "values")[1]

        popup = Toplevel(self.root)
        popup.title(f"{self._('note_popup_title')}: {file_name}")

        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        w = max(400, int(sw * 0.3))
        h = max(300, int(sh * 0.4))
        x, y = (sw - w) // 2, (sh - h) // 2
        popup.geometry(f"{w}x{h}+{x}+{y}")

        popup.transient(self.root)
        popup.grab_set()

        main_frame = ttk.Frame(popup, padding=10)
        main_frame.pack(fill="both", expand=True)
        
        note_text = tk.Text(main_frame, wrap="word", height=10)
        note_text.pack(fill="both", expand=True, pady=(0, 10))
        
        # It looks up the note for this specific path.
        existing_note = self.file_annotations.get(path_str, "")
        if existing_note:
            note_text.insert("1.0", existing_note)

        def save_note():
            new_note = note_text.get("1.0", tk.END).strip()
            # The new note is saved using the captured path_str as the unique key.
            # This ensures the note is only associated with the correct file.
            if new_note:
                self.file_annotations[path_str] = new_note
            elif path_str in self.file_annotations:
                del self.file_annotations[path_str]
            
            self.dirty_notes.add(path_str)
            self.case_is_dirty = True
            if self.is_reader_mode:
                self.file_menu.entryconfig(self._("menu_save_case_simple"), state="normal")
            
            self._apply_filter() 
            
            new_item_to_select = None
            for child_id in self.tree.get_children(""):
                if self.tree.item(child_id, "values")[4] == path_str:
                    new_item_to_select = child_id
                    break

            if new_item_to_select:
                self.tree.selection_set(new_item_to_select)
                self.tree.focus(new_item_to_select)
                self.on_select_item(None)

            popup.destroy()

        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x")
        
        save_button = ttk.Button(button_frame, text=self._("settings_save"), command=save_note)
        save_button.pack(side="right", padx=5)
        
        cancel_button = ttk.Button(button_frame, text=self._("settings_cancel"), command=popup.destroy)
        cancel_button.pack(side="right")
        
       
    def _setup_drag_and_drop(self):
        """Enables drag and drop for the main window, unless in reader mode."""
        if self.is_reader_mode:
            return
        self.root.drop_target_register(DND_FILES)
        self.root.dnd_bind('<<Drop>>', self.handle_drop)

    def _save_current_case(self):
        """Saves annotations back to the currently open .prc file (for Reader mode)."""
        if not self.current_case_filepath or not self.case_is_dirty:
            return
        
        try:
            self._write_case_to_file(self.current_case_filepath)
            self.case_is_dirty = False
            self.dirty_notes.clear()
            if self.is_reader_mode:
                self.file_menu.entryconfig(self._("menu_save_case_simple"), state="disabled")
            logging.info(f"Annotations saved to case file: {self.current_case_filepath}")

            # Refresh the GUI to reflect the saved state
            self._apply_filter()

        except Exception as e:
            logging.error(f"Failed to save case file '{self.current_case_filepath}': {e}")
            messagebox.showerror(self._("case_save_error_title"), self._("case_save_error_msg").format(e=e))
            
    def handle_drop(self, event):
        """Handles files that are dropped onto the window."""
        # The path can sometimes be enclosed in braces {}
        folder_path = event.data.strip('{}')
        if os.path.isdir(folder_path):
            self.start_scan_thread(Path(folder_path))
        else:
            messagebox.showwarning(self._("drop_error_title"), self._("drop_error_message"))

    def _on_tree_motion(self, event):
        """Changes the cursor to a hand when hovering over a clickable cell."""
        col_id = self.tree.identify_column(event.x)
        row_id = self.tree.identify_row(event.y)
        if not row_id:
            self.tree.config(cursor="")
            return

        path_str = self.tree.item(row_id, "values")[3]
        
        # Check for EXIFTool column (index 8)
        if col_id == '#8':
            if path_str in self.exif_outputs and self.exif_outputs[path_str]:
                exif_output = self.exif_outputs[path_str]
                # Check if the output is an error message
                is_error = (exif_output == self._("exif_err_notfound") or
                            exif_output.startswith(self._("exif_err_prefix")) or
                            exif_output.startswith(self._("exif_err_run").split("{")[0]))
                if not is_error:
                    self.tree.config(cursor="hand2")
                    return
        
        # Check for Indicators column (index 9)
        if col_id == '#9':
            data_item = next((d for d in self.all_scan_data if str(d.get('path')) == path_str), None)
            if data_item and data_item.get("indicator_keys"):
                self.tree.config(cursor="hand2")
                return

        # Default cursor if not over a clickable cell
        self.tree.config(cursor="")
    
    def show_inspector_popup(self, event=None):
        """
        Shows and updates a single, reusable inspector window with details of the selected file,
        including an integrated PDF viewer.
        """
        # --- Identify the selected item ---
        item_id = None
        if event: # Called from a double-click
            if self.tree.identify_region(event.x, event.y) == "heading":
                return
            item_id = self.tree.identify_row(event.y)
        else: # Called from a menu or selection change
            selected_items = self.tree.selection()
            if selected_items:
                item_id = selected_items[0]

        if not item_id:
            return

        values = self.tree.item(item_id, "values")
        path_str = values[4] 
        file_name = values[1]
        resolved_path = self._resolve_case_path(path_str)
        file_data = next((d for d in self.all_scan_data if str(d.get('path')) == path_str), None)
        if not file_data:
            return

        # --- Create the Inspector window once if it doesn't exist ---
        if not self.inspector_window or not self.inspector_window.winfo_exists():
            self.inspector_window = Toplevel(self.root)
            self.inspector_window.title("Inspector")
            
            sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
            w, h = int(sw * 0.75), int(sh * 0.8)
            x, y = (sw - w) // 2, (sh - h) // 2
            self.inspector_window.geometry(f"{w}x{h}+{x}+{y}")

            notebook = ttk.Notebook(self.inspector_window)
            notebook.pack(pady=10, padx=10, fill="both", expand=True)

            # Tab 1: Details
            indicators_frame = ttk.Frame(notebook, padding="10")
            notebook.add(indicators_frame, text=self._("inspector_details_tab"))
            self.inspector_indicators_text = tk.Text(indicators_frame, wrap="word", font=("Segoe UI", 9))
            self.inspector_indicators_text.pack(fill="both", expand=True)
            self.inspector_indicators_text.tag_configure("bold", font=("Segoe UI", 9, "bold"))
            self._make_text_copyable(self.inspector_indicators_text)

            # Tab 2: EXIFTool
            exif_frame = ttk.Frame(notebook, padding="10")
            notebook.add(exif_frame, text=self._("col_exif"))
            exif_text_widget = tk.Text(exif_frame, wrap="word", font=("Consolas", 10))
            exif_vscroll = ttk.Scrollbar(exif_frame, orient="vertical", command=exif_text_widget.yview)
            exif_text_widget.config(yscrollcommand=exif_vscroll.set)
            exif_vscroll.pack(side="right", fill="y")
            exif_text_widget.pack(fill="both", expand=True)
            self.inspector_exif_text = exif_text_widget
            self._make_text_copyable(self.inspector_exif_text)

            # Tab 3: Timeline
            timeline_frame = ttk.Frame(notebook, padding="10")
            notebook.add(timeline_frame, text=self._("show_timeline"))
            timeline_text_widget = tk.Text(timeline_frame, wrap="word", font=("Courier New", 10))
            timeline_vscroll = ttk.Scrollbar(timeline_frame, orient="vertical", command=timeline_text_widget.yview)
            timeline_text_widget.config(yscrollcommand=timeline_vscroll.set)
            timeline_vscroll.pack(side="right", fill="y")
            timeline_text_widget.pack(fill="both", expand=True)
            self.inspector_timeline_text = timeline_text_widget
            self._make_text_copyable(self.inspector_timeline_text)

            # Tab 4: PDF Viewer
            pdf_view_frame = ttk.Frame(notebook)
            notebook.add(pdf_view_frame, text=self._("view_pdf"))
            self.inspector_pdf_frame = pdf_view_frame
            
            def on_inspector_close():
                if self.inspector_doc:
                    self.inspector_doc.close()
                    self.inspector_doc = None
                self.inspector_window.withdraw()
            
            self.inspector_window.protocol("WM_DELETE_WINDOW", on_inspector_close)

        # --- Update the content of the existing window ---
        self.inspector_window.title(f"Inspector: {file_name}")

        # Update Details Tab
        self.inspector_indicators_text.config(state="normal")
        self.inspector_indicators_text.delete("1.0", tk.END)
        for i, val in enumerate(values):
            col_name = self.tree.heading(self.columns[i], "text")
            self.inspector_indicators_text.insert(tk.END, f"{col_name}: ", ("bold",))
            if col_name == self._("col_indicators") and file_data and file_data.get("indicator_keys"):
                indicator_details = [self._format_indicator_details(key, details) for key, details in file_data["indicator_keys"].items()]
                full_indicators_str = "\n  • " + "\n  • ".join(indicator_details)
                self.inspector_indicators_text.insert(tk.END, full_indicators_str + "\n")
            else:
                self.inspector_indicators_text.insert(tk.END, val + "\n")
        note = self.file_annotations.get(path_str)
        if note:
            self.inspector_indicators_text.insert(tk.END, "\n" + "-"*40 + "\n")
            self.inspector_indicators_text.insert(tk.END, f"{self._('note_label')}\n", ("bold",))
            self.inspector_indicators_text.insert(tk.END, note)
        self.inspector_indicators_text.config(state="disabled")

        # Update EXIF Tab
        self.inspector_exif_text.config(state="normal")
        self.inspector_exif_text.delete("1.0", tk.END)
        self.inspector_exif_text.insert("1.0", self.exif_outputs.get(path_str, self._("no_exif_output_message")))
        self.inspector_exif_text.config(state="disabled")

        # Update Timeline Tab
        self.inspector_timeline_text.config(state="normal")
        self.inspector_timeline_text.delete("1.0", tk.END)
        self._populate_timeline_widget(self.inspector_timeline_text, path_str)
        self.inspector_timeline_text.config(state="disabled")

        # --- PDF Viewer Update Logic with crash fix ---
        if self.inspector_pdf_update_job:
            self.inspector_window.after_cancel(self.inspector_pdf_update_job)
            self.inspector_pdf_update_job = None

        for widget in self.inspector_pdf_frame.winfo_children():
            widget.destroy()
        
        if self.inspector_doc:
            self.inspector_doc.close()
        try:
            self.inspector_doc = fitz.open(resolved_path)
        except Exception:
            self.inspector_doc = None

        if self.inspector_doc:
            pdf_main_frame = ttk.Frame(self.inspector_pdf_frame, padding=10)
            pdf_main_frame.pack(fill="both", expand=True)
            pdf_main_frame.rowconfigure(0, weight=1)
            pdf_main_frame.columnconfigure(0, weight=1)
            pdf_image_label = ttk.Label(pdf_main_frame)
            pdf_image_label.grid(row=0, column=0, pady=5, sticky="nsew")
            pdf_nav_frame = ttk.Frame(pdf_main_frame)
            pdf_nav_frame.grid(row=1, column=0, pady=(10,0))
            
            prev_button = ttk.Button(pdf_nav_frame, text=self._("diff_prev_page"))
            page_label = ttk.Label(pdf_nav_frame, text="", font=("Segoe UI", 9, "italic"))
            next_button = ttk.Button(pdf_nav_frame, text=self._("diff_next_page"))
            prev_button.pack(side="left", padx=10)
            page_label.pack(side="left", padx=10)
            next_button.pack(side="left", padx=10)

            current_page_ref = {'page': 0}
            total_pages = len(self.inspector_doc)

            def update_page(page_num):
                if not (0 <= page_num < total_pages): return
                current_page_ref['page'] = page_num
                
                try:
                    if not self.inspector_window.winfo_exists() or not pdf_main_frame.winfo_exists():
                        return
                except tk.TclError:
                    return

                page = self.inspector_doc.load_page(page_num)
                pix = page.get_pixmap(dpi=150)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                
                if pdf_main_frame.winfo_width() <= 1 or pdf_main_frame.winfo_height() <= 1:
                    self.inspector_pdf_update_job = self.inspector_window.after(50, lambda: update_page(page_num))
                    return
                
                max_w = pdf_main_frame.winfo_width() * 0.95
                max_h = pdf_main_frame.winfo_height() * 0.95
                ratio = min(max_w / img.width, max_h / img.height) if img.width > 0 and img.height > 0 else 1
                scaled_size = (int(img.width * ratio), int(img.height * ratio))

                img_tk = ImageTk.PhotoImage(img.resize(scaled_size, Image.Resampling.LANCZOS))
                pdf_image_label.img_tk = img_tk
                
                pdf_image_label.config(image=img_tk)
                page_label.config(text=self._("diff_page_label").format(current=page_num + 1, total=total_pages))
                prev_button.config(state="normal" if page_num > 0 else "disabled")
                next_button.config(state="normal" if page_num < total_pages - 1 else "disabled")

            prev_button.config(command=lambda: update_page(current_page_ref['page'] - 1))
            next_button.config(command=lambda: update_page(current_page_ref['page'] + 1))
            
            self.inspector_pdf_update_job = self.inspector_window.after(100, lambda: update_page(0))
        else:
            ttk.Label(self.inspector_pdf_frame, text="Could not display PDF.").pack(pady=20)

        # --- Show and raise the window ---
        self.inspector_window.deiconify()
        self.inspector_window.lift()

     
    def show_context_menu(self, event):
        """Displays a simplified right-click context menu for the selected row."""
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return
        
        self.tree.selection_set(item_id)
        values = self.tree.item(item_id, "values")
        path_str = values[4] if values else None
        file_data = next((d for d in self.all_scan_data if str(d.get('path')) == path_str), None)

        context_menu = tk.Menu(self.root, tearoff=0)
        
        # Main actions
        context_menu.add_command(label="Inspector...", command=self.show_inspector_popup)
        context_menu.add_command(label=self._("menu_add_note"), command=self._show_note_popup)
        context_menu.add_separator()
        
        # Conditional actions
        text_diff_available = file_data and file_data.get("indicator_keys", {}).get("TouchUp_TextEdit", {}).get("text_diff")
        if text_diff_available:
            context_menu.add_command(label="View Text Diff", command=lambda: self.show_text_diff_popup(item_id))
        
        is_revision = file_data and file_data.get('is_revision')
        if is_revision:
            context_menu.add_command(label=self._("visual_diff"), command=lambda: self.show_visual_diff_popup(item_id))

        context_menu.add_separator()
        context_menu.add_command(label="Open File Location", command=lambda: self.open_file_location(item_id))
        
        context_menu.tk_popup(event.x_root, event.y_root)  
    
    def show_text_diff_popup(self, item_id):
        """Displays a popup showing the text differences between a file and its revision."""
        path_str = self.tree.item(item_id, "values")[4]
        file_data = next((d for d in self.all_scan_data if str(d.get('path')) == path_str), None)
        
        text_diff_data = file_data.get("indicator_keys", {}).get("TouchUp_TextEdit", {}).get("text_diff")
        if not text_diff_data:
            messagebox.showinfo("No Diff", "No text comparison data is available for this file.", parent=self.root)
            return

        popup = Toplevel(self.root)
        popup.title(f"Text Comparison for {file_data['path'].name}")
        
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        w, h = int(sw * 0.7), int(sh * 0.75)
        x, y = (sw - w) // 2, (sh - h) // 2
        popup.geometry(f"{w}x{h}+{x}+{y}")
        
        popup.transient(self.root)

        frame = ttk.Frame(popup, padding=10)
        frame.pack(fill="both", expand=True)
        
        text_widget = tk.Text(frame, wrap="word", font=("Courier New", 10))
        v_scroll = ttk.Scrollbar(frame, orient="vertical", command=text_widget.yview)
        text_widget.config(yscrollcommand=v_scroll.set)
        v_scroll.pack(side="right", fill="y")
        text_widget.pack(side="left", fill="both", expand=True)

        text_widget.tag_configure("addition", foreground="green")
        text_widget.tag_configure("deletion", foreground="red")

        for line in text_diff_data:
            if line.startswith('+ '):
                text_widget.insert(tk.END, line, "addition")
            elif line.startswith('- '):
                text_widget.insert(tk.END, line, "deletion")
            else:
                text_widget.insert(tk.END, line)
        
        self._make_text_copyable(text_widget)
        
    
    def open_file_location(self, item_id):
        """Opens the folder containing the selected file in the system's file explorer."""
        values = self.tree.item(item_id, "values")
        if values:
            path_str = values[4]
            resolved_path = self._resolve_case_path(path_str)
            if resolved_path and resolved_path.exists():
                webbrowser.open(os.path.dirname(resolved_path))
            else:
                messagebox.showwarning("File Not Found", f"The file could not be found at the path:\n{resolved_path}")       

    def _make_text_copyable(self, text_widget):
        """Makes a Text widget read-only but allows text selection and copying."""
        context_menu = tk.Menu(text_widget, tearoff=0)
        
        def copy_selection(event=None):
            """Copies the selected text to the clipboard."""
            try:
                selected_text = text_widget.get(tk.SEL_FIRST, tk.SEL_LAST)
                self.root.clipboard_clear()
                self.root.clipboard_append(selected_text)
            except tk.TclError:
                # This can happen if no text is selected
                pass
            return "break" # Prevents default event handling

        context_menu.add_command(label=self._("copy"), command=copy_selection)
        

        def show_context_menu(event):
            """Shows the context menu if text is selected."""
            if text_widget.tag_ranges(tk.SEL):
                context_menu.tk_popup(event.x_root, event.y_root)

        text_widget.config(state="normal") # Make it temporarily writable to bind
        text_widget.bind("<Key>", lambda e: "break") # Disable typing
        text_widget.bind("<Button-3>", show_context_menu) # Right-click
        text_widget.bind("<Control-c>", copy_selection) # Ctrl+C
        text_widget.bind("<Command-c>", copy_selection) # Command+C for macOS
        
    def _add_layer_indicators(self, raw: bytes, path: Path, indicators: dict):
        """
        Adds indicators for layers:
          - "Has Layers (count)" if OCGs are found.
          - "More Layers Than Pages" if layer count > page count.
        """
        try:
            layers_cnt = count_layers(raw)
        except Exception:
            layers_cnt = 0

        if layers_cnt <= 0:
            return

        indicators['HasLayers'] = {'count': layers_cnt}

        # Compare with page count (best-effort)
        page_count = 0
        try:
            with fitz.open(path) as _doc:
                page_count = _doc.page_count
        except Exception:
            pass

        if page_count and layers_cnt > page_count:
            indicators['MoreLayersThanPages'] = {'layers': layers_cnt, 'pages': page_count}

    def show_visual_diff_popup(self, item_id):
        """Shows a side-by-side visual comparison of a revision and its original."""
        self.root.config(cursor="watch")
        self.root.update_idletasks()

        rev_path_str = self.tree.item(item_id, "values")[4]
        original_path_str = next((d['original_path'] for d in self.all_scan_data if str(d['path']) == rev_path_str), None)

        if not original_path_str:
            messagebox.showerror(self._("diff_error_title"), "Original file for revision not found.", parent=self.root)
            self.root.config(cursor="")
            return

        resolved_rev_path = self._resolve_case_path(rev_path_str)
        resolved_orig_path = self._resolve_case_path(original_path_str)

        try:
            # --- Popup Window Setup ---
            popup = Toplevel(self.root)
            popup.title(self._("diff_popup_title"))
            
            popup.current_page = 0
            popup.path_orig = resolved_orig_path
            popup.path_rev = resolved_rev_path
            with fitz.open(popup.path_orig) as doc:
                popup.total_pages = doc.page_count

            # --- Widget Layout ---
            main_frame = ttk.Frame(popup, padding=10)
            main_frame.pack(fill="both", expand=True)

            image_frame = ttk.Frame(main_frame)
            image_frame.grid(row=1, column=0, columnspan=3, pady=5)
            
            label_orig = ttk.Label(image_frame)
            label_rev = ttk.Label(image_frame)
            label_diff = ttk.Label(image_frame)
            label_orig.grid(row=1, column=0, padx=5)
            label_rev.grid(row=1, column=1, padx=5)
            label_diff.grid(row=1, column=2, padx=5)

            ttk.Label(image_frame, text=self._("diff_original_label"), font=("Segoe UI", 10, "bold")).grid(row=0, column=0)
            ttk.Label(image_frame, text=self._("diff_revision_label"), font=("Segoe UI", 10, "bold")).grid(row=0, column=1)
            ttk.Label(image_frame, text=self._("diff_differences_label"), font=("Segoe UI", 10, "bold")).grid(row=0, column=2)

            nav_frame = ttk.Frame(main_frame)
            nav_frame.grid(row=2, column=0, columnspan=3, pady=(10,0))
            
            prev_button = ttk.Button(nav_frame, text=self._("diff_prev_page"))
            page_label = ttk.Label(nav_frame, text="", font=("Segoe UI", 9, "italic"))
            next_button = ttk.Button(nav_frame, text=self._("diff_next_page"))

            prev_button.pack(side="left", padx=10)
            page_label.pack(side="left", padx=10)
            next_button.pack(side="left", padx=10)
            
            def update_page(page_num):
                """Renders the original, revision, and difference for a specific page."""
                if not (0 <= page_num < popup.total_pages):
                    return
                
                popup.current_page = page_num
                self.root.config(cursor="watch")
                self.root.update()

                with fitz.open(popup.path_orig) as doc_orig, fitz.open(popup.path_rev) as doc_rev:
                    page_orig = doc_orig.load_page(page_num)
                    page_rev = doc_rev.load_page(page_num)

                    pix_orig = page_orig.get_pixmap(dpi=150)
                    pix_rev = page_rev.get_pixmap(dpi=150)
                
                img_orig = Image.frombytes("RGB", [pix_orig.width, pix_orig.height], pix_orig.samples)
                img_rev = Image.frombytes("RGB", [pix_rev.width, pix_rev.height], pix_rev.samples)

                diff = ImageChops.difference(img_orig, img_rev)
                mask = diff.convert('L').point(lambda x: 255 if x > 20 else 0)
                final_diff = Image.composite(Image.new('RGB', img_orig.size, 'red'), ImageOps.grayscale(img_orig).convert('RGB'), mask)
                
                screen_w, screen_h = popup.winfo_screenwidth(), popup.winfo_screenheight()
                max_img_w, max_img_h = (screen_w * 0.95) / 3, screen_h * 0.8
                orig_w, orig_h = img_orig.size
                ratio = min(max_img_w / orig_w, max_img_h / orig_h) if orig_w > 0 and orig_h > 0 else 1
                scaled_size = (int(orig_w * ratio), int(orig_h * ratio))

                images_tk = [ImageTk.PhotoImage(img.resize(scaled_size, Image.Resampling.LANCZOS)) for img in [img_orig, img_rev, final_diff]]
                popup.images_tk = images_tk
                
                label_orig.config(image=images_tk[0])
                label_rev.config(image=images_tk[1])
                label_diff.config(image=images_tk[2])

                page_label.config(text=self._("diff_page_label").format(current=page_num + 1, total=popup.total_pages))
                prev_button.config(state="normal" if page_num > 0 else "disabled")
                next_button.config(state="normal" if page_num < popup.total_pages - 1 else "disabled")
                self.root.config(cursor="")

            prev_button.config(command=lambda: update_page(popup.current_page - 1))
            next_button.config(command=lambda: update_page(popup.current_page + 1))

            update_page(0)
            
            popup.transient(self.root)
            popup.grab_set()

        except Exception as e:
            logging.error(f"Visual diff error: {e}")
            messagebox.showerror(self._("diff_error_title"), self._("diff_error_msg").format(e=e), parent=self.root)
            self.root.config(cursor="")
    
    def open_settings_popup(self):
        """Opens a window to edit application settings from config.ini."""
        settings_popup = Toplevel(self.root)
        settings_popup.title(self._("settings_title"))
        settings_popup.transient(self.root)
        settings_popup.geometry("400x230")
        settings_popup.resizable(False, False)

        main_frame = ttk.Frame(settings_popup, padding=15)
        main_frame.pack(expand=True, fill="both")

        # Create StringVars to hold the values from the entry boxes
        size_var = tk.StringVar(value=str(PDFReconConfig.MAX_FILE_SIZE // (1024*1024)))
        timeout_var = tk.StringVar(value=str(PDFReconConfig.EXIFTOOL_TIMEOUT))
        threads_var = tk.StringVar(value=str(PDFReconConfig.MAX_WORKER_THREADS))
        diff_pages_var = tk.StringVar(value=str(PDFReconConfig.VISUAL_DIFF_PAGE_LIMIT))
        export_xref_var = tk.BooleanVar(value=PDFReconConfig.EXPORT_INVALID_XREF)

        # Create labels and entry fields
        fields = [
            (self._("settings_max_size"), size_var),
            (self._("settings_timeout"), timeout_var),
            (self._("settings_threads"), threads_var),
            (self._("settings_diff_pages"), diff_pages_var),
        ]

        for i, (label_text, var) in enumerate(fields):
            label = ttk.Label(main_frame, text=label_text)
            label.grid(row=i, column=0, sticky="w", pady=5)
            entry = ttk.Entry(main_frame, textvariable=var, width=10)
            entry.grid(row=i, column=1, sticky="e", pady=5)
        
        main_frame.columnconfigure(1, weight=1)
        xref_check = ttk.Checkbutton(main_frame, text=self._("settings_export_invalid_xref"), variable=export_xref_var)
        xref_check.grid(row=len(fields), column=0, columnspan=2, sticky="w", pady=5)

        def save_settings():
            """Validates, saves, and applies the new settings."""
            try:
                # Read and validate values
                new_size = int(size_var.get())
                new_timeout = int(timeout_var.get())
                new_threads = int(threads_var.get())
                new_diff_pages = int(diff_pages_var.get())
                new_export_xref = export_xref_var.get()

                # Update the running config
                PDFReconConfig.MAX_FILE_SIZE = new_size * 1024 * 1024
                PDFReconConfig.EXIFTOOL_TIMEOUT = new_timeout
                PDFReconConfig.MAX_WORKER_THREADS = new_threads
                PDFReconConfig.VISUAL_DIFF_PAGE_LIMIT = new_diff_pages
                PDFReconConfig.EXPORT_INVALID_XREF = new_export_xref

                # Write to config.ini
                parser = configparser.ConfigParser()
                parser.read(self.config_path)
                if 'Settings' not in parser:
                    parser['Settings'] = {}
                
                parser['Settings']['MaxFileSizeMB'] = str(new_size)
                parser['Settings']['ExifToolTimeout'] = str(new_timeout)
                parser['Settings']['MaxWorkerThreads'] = str(new_threads)
                parser['Settings']['VisualDiffPageLimit'] = str(new_diff_pages)
                parser['Settings']['ExportInvalidXREF'] = str(new_export_xref)

                with open(self.config_path, 'w') as configfile:
                    configfile.write("# PDFRecon Configuration File\n")
                    parser.write(configfile)

                logging.info("Settings updated and saved to config.ini.")
                messagebox.showinfo(self._("settings_saved_title"), self._("settings_saved_msg"), parent=settings_popup)
                settings_popup.destroy()

            except ValueError:
                messagebox.showerror("Error", self._("settings_invalid_input"), parent=settings_popup)
            except Exception as e:
                messagebox.showerror("Error", f"Could not save settings: {e}", parent=settings_popup)
                logging.error(f"Failed to save settings: {e}")

        # --- Buttons Frame ---
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.grid(row=len(fields) + 1, column=0, columnspan=2, pady=(15, 0))
        
        save_button = ttk.Button(buttons_frame, text=self._("settings_save"), command=save_settings)
        save_button.pack(side="left", padx=5)

        cancel_button = ttk.Button(buttons_frame, text=self._("settings_cancel"), command=settings_popup.destroy)
        cancel_button.pack(side="left", padx=5)

        settings_popup.grab_set()
        self.root.wait_window(settings_popup)


        
    def _resolve_path(self, filename, base_is_parent=False):
        """
        Resolves the correct path for a resource file, whether running as a script
        or as a frozen executable (e.g., with PyInstaller).
        """
        if getattr(sys, 'frozen', False):
            # If the app is frozen, the base path is the folder containing the exe.
            base_path = Path(sys.executable).parent
            if not base_is_parent:
                # Data files (like exiftool) are in a temp folder _MEIPASS,
                # unless we want a file next to the exe (base_is_parent=True).
                return Path(getattr(sys, '_MEIPASS', base_path)) / filename
        else:
            # If running as a normal script, the base path is the script's folder.
            base_path = Path(__file__).resolve().parent
        return base_path / filename

    def show_license(self):
        """Displays the license information from 'license.txt' in a popup."""
        license_path = self._resolve_path("license.txt")
        try:
            with open(license_path, 'r', encoding='utf-8') as f: license_text = f.read()
        except FileNotFoundError:
            messagebox.showerror(self._("license_error_title"), self._("license_error_message"))
            return
        
        # --- Popup Window Setup ---
        license_popup = Toplevel(self.root)
        license_popup.title(self._("license_popup_title"))
        license_popup.geometry("600x500")
        license_popup.transient(self.root)

        # --- Text Widget with Scrollbar and Close Button ---
        text_frame = ttk.Frame(license_popup, padding=10)
        text_frame.pack(fill="both", expand=True)
        text_frame.rowconfigure(0, weight=1)
        text_frame.columnconfigure(0, weight=1)
        
        scroll_frame = ttk.Frame(text_frame)
        scroll_frame.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(scroll_frame)
        scrollbar.pack(side="right", fill="y")
        text_widget = tk.Text(scroll_frame, wrap="word", yscrollcommand=scrollbar.set, font=("Courier New", 9), borderwidth=0, highlightthickness=0)
        text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=text_widget.yview)
        
        text_widget.insert("1.0", license_text)
        text_widget.config(state="disabled") # Make read-only
        
        close_button = ttk.Button(text_frame, text=self._("close_button_text"), command=license_popup.destroy)
        close_button.grid(row=1, column=0, pady=(10,0))
        
    def show_log_file(self):
        """Opens the application's log file in the default system viewer."""
        if self.log_file_path.exists():
            webbrowser.open(self.log_file_path.as_uri())
        else:
            messagebox.showinfo(self._("log_not_found_title"), self._("log_not_found_message"), parent=self.root)

    def _sort_column(self, col, reverse):
        """Sorts the treeview column when its header is clicked."""
        is_id_column = col == self.columns[0]
        # Define a key for sorting: convert to int for ID column, otherwise use string value
        def get_key(item):
            val = self.tree.set(item, col)
            return int(val) if is_id_column and val else val

        # Get data from tree, sort it, and re-insert it
        data_list = [(get_key(k), k) for k in self.tree.get_children("")]
        data_list.sort(reverse=reverse)
        for index, (val, k) in enumerate(data_list):
            self.tree.move(k, "", index)
        
        # Toggle the sort direction for the next click
        self.tree.heading(col, command=lambda: self._sort_column(col, not reverse))

    def choose_folder(self):
        """Opens a dialog for the user to select a folder to scan."""
        folder_path = filedialog.askdirectory(title=self._("choose_folder_title"))
        if folder_path:
            self.start_scan_thread(Path(folder_path))

    def start_scan_thread(self, folder_path):
        """Initializes and starts the background scanning process."""
        logging.info(f"Starting scan of folder: {folder_path}")
        
        # --- Reset Application State ---
        self._reset_state()
        self.last_scan_folder = folder_path
        self.case_root_path = folder_path # Set the root for the new scan
        
        # --- Update GUI for Scanning State ---
        self.scan_button.config(state="disabled")
        self.export_menubutton.config(state="disabled")
        if not self.is_reader_mode:
            self.file_menu.entryconfig(self._("menu_save_case"), state="disabled")
            if getattr(sys, 'frozen', False):
                 self.file_menu.entryconfig(self._("menu_export_reader"), state="disabled")

        self.status_var.set(self._("preparing_analysis"))
        self.progressbar.config(value=0)

        # --- Create a dedicated thread pool for copy operations ---
        self.copy_executor = ThreadPoolExecutor(max_workers=4, thread_name_prefix='CopyWorker')

        # --- Start Main Worker Thread ---
        scan_thread = threading.Thread(target=self._scan_worker_parallel, args=(folder_path, self.scan_queue))
        scan_thread.daemon = True
        scan_thread.start()

        self._process_queue()
       
    def _reset_state(self):
        """Resets all data and GUI elements to their initial state."""
        self.tree.delete(*self.tree.get_children())
        self.report_data.clear()
        self.all_scan_data.clear()
        self.exif_outputs.clear()
        self.timeline_data.clear()
        self.path_to_id.clear()
        self.evidence_hashes.clear()
        self.revision_counter = 0
        self.scan_queue = queue.Queue()
        self.scan_start_time = time.time()
        self.filter_var.set("")
        self.last_scan_folder = None
        self.current_case_filepath = None
        self.case_is_dirty = False
        self.dirty_notes.clear()
        self.detail_text.config(state="normal")
        self.detail_text.delete("1.0", tk.END)
        self.detail_text.config(state="disabled")

    def _open_case(self, filepath=None):
        """Opens a dialog to load a case, or loads a case from a given filepath."""
        if not filepath:
            if self.all_scan_data:
                if not messagebox.askokcancel(self._("case_open_warning_title"), self._("case_open_warning_msg")):
                    return
            
            filepath = filedialog.askopenfilename(
                title="Open PDFRecon Case",
                filetypes=[("PDFRecon Case Files", "*.prc"), ("All files", "*.*")]
            )
        
        if not filepath:
            return

        try:
            with open(filepath, 'rb') as f:
                case_data = pickle.load(f)

            # --- Restore State from Case File ---
            self._reset_state()
            self.current_case_filepath = filepath # Store the path
            self.case_is_dirty = False
            self._update_title() # Update title
            self.case_root_path = Path(filepath).parent 
            self.all_scan_data = case_data.get('all_scan_data', [])
            self.file_annotations = case_data.get('file_annotations', {})
            self.exif_outputs = case_data.get('exif_outputs', {})
            self.dirty_notes.clear()
            self.timeline_data = case_data.get('timeline_data', {})
            self.path_to_id = case_data.get('path_to_id', {})
            self.evidence_hashes = case_data.get('evidence_hashes', {})
            self.revision_counter = case_data.get('revision_counter', 0)
            self.last_scan_folder = case_data.get('scan_folder', None)
            
            # --- Update GUI with loaded data ---
            self._apply_filter()
            self._update_summary_status()
            self.export_menubutton.config(state="normal")
            
            if self.evidence_hashes:
                self.file_menu.entryconfig(self._("menu_verify_integrity"), state="normal")

            if not self.is_reader_mode:
                self.file_menu.entryconfig(self._("menu_save_case"), state="normal")
                if getattr(sys, 'frozen', False):
                    self.file_menu.entryconfig(self._("menu_export_reader"), state="normal")

            logging.info(f"Successfully loaded case file: {filepath}")

        except Exception as e:
            logging.error(f"Failed to open case file '{filepath}': {e}")
            messagebox.showerror(self._("case_open_error_title"), self._("case_open_error_msg").format(e=e))
            
    def _write_case_to_file(self, filepath):
        """Helper function to gather and write case data to a specific file path."""
        case_data = {
            'app_version': self.app_version,
            'scan_folder': self.last_scan_folder,
            'all_scan_data': self.all_scan_data,
            'file_annotations': self.file_annotations,
            'exif_outputs': self.exif_outputs,
            'timeline_data': self.timeline_data,
            'path_to_id': self.path_to_id,
            'revision_counter': self.revision_counter,
            'evidence_hashes': self.evidence_hashes,
        }
        with open(filepath, 'wb') as f:
            pickle.dump(case_data, f)
    
    def _hash_file(self, filepath):
        """Calculates the SHA-256 hash of a file."""
        sha256_hash = hashlib.sha256()
        try:
            with open(filepath, "rb") as f:
                for byte_block in iter(lambda: f.read(4096), b""):
                    sha256_hash.update(byte_block)
            return sha256_hash.hexdigest()
        except FileNotFoundError:
            logging.error(f"Could not hash file, not found: {filepath}")
            return None
        except Exception as e:
            logging.error(f"Error hashing file {filepath}: {e}")
            return None

    def _calculate_hashes(self, data_to_hash):
        """Calculates hashes for all files listed in the data and returns a dictionary."""
        hashes = {}
        for item in data_to_hash:
            path_str = item.get('path')
            if path_str:
                full_path = self._resolve_case_path(path_str)
                file_hash = self._hash_file(full_path)
                if file_hash:
                    hashes[str(path_str)] = file_hash
        return hashes

    def _verify_integrity(self):
        """Verifies the integrity of all evidence files against the stored hashes and logs the result."""
        if not self.evidence_hashes:
            messagebox.showinfo(self._("verify_title"), self._("verify_no_hashes"))
            return

        self.status_var.set(self._("verify_running"))
        self.root.update_idletasks()

        mismatched_files = []
        missing_files = []
        
        total_files = len(self.evidence_hashes)
        verified_count = 0

        for path_str, original_hash in self.evidence_hashes.items():
            full_path = self._resolve_case_path(path_str)
            if not full_path or not full_path.exists():
                missing_files.append(str(path_str))
                continue
            
            current_hash = self._hash_file(full_path)
            if current_hash != original_hash:
                mismatched_files.append(str(path_str))
            
            verified_count += 1

        self._update_summary_status()

        if not mismatched_files and not missing_files:
            logging.info(f"Integrity check result: Success. All {verified_count}/{total_files} files are valid.")
            messagebox.showinfo(self._("verify_fail_title"), self._("verify_success"))
        else:
            log_summary = (f"Integrity check result: FAILURE. "
                           f"Verified: {verified_count}/{total_files}, "
                           f"Mismatched: {len(mismatched_files)}, "
                           f"Missing: {len(missing_files)}")
            logging.warning(log_summary)

            report_lines = []
            report_popup = Toplevel(self.root)
            report_popup.title(self._("verify_fail_title"))
            
            sw = self.root.winfo_screenwidth()
            sh = self.root.winfo_screenheight()
            w = max(700, int(sw * 0.5))
            h = max(450, int(sh * 0.6))
            x, y = (sw - w) // 2, (sh - h) // 2
            report_popup.geometry(f"{w}x{h}+{x}+{y}")
            
            text_frame = ttk.Frame(report_popup, padding=10)
            text_frame.pack(fill="both", expand=True)
            text_widget = tk.Text(text_frame, wrap="word", font=("Courier New", 9))
            text_widget.pack(fill="both", expand=True)

            def add_line(text):
                report_lines.append(text)
                text_widget.insert(tk.END, text + "\n")

            add_line(f"{self._('verify_report_header')}")
            add_line("-----------------------------")
            add_line(f"{self._('verify_report_verified')}: {verified_count}/{total_files}")
            add_line(f"{self._('verify_report_mismatched')}: {len(mismatched_files)}")
            add_line(f"{self._('verify_report_missing')}: {len(missing_files)}\n")

            if mismatched_files:
                add_line(f"{self._('verify_report_modified_header')}")
                logging.warning("Mismatched files:")
                for f in mismatched_files:
                    add_line(f"- {f}")
                    logging.warning(f"- {f}")
                add_line("")
            
            if missing_files:
                add_line(f"{self._('verify_report_missing_header')}")
                logging.warning("Missing files:")
                for f in missing_files:
                    add_line(f"- {f}")
                    logging.warning(f"- {f}")

            text_widget.config(state="disabled")
            messagebox.showwarning(self._("verify_fail_title"), self._("verify_fail_msg"), parent=report_popup)
            
    def _save_case(self):
        """Saves the current analysis results to a case file."""
        if not self.all_scan_data:
            messagebox.showwarning(self._("case_nothing_to_save_title"), self._("case_nothing_to_save_msg"))
            return
            
        filepath = filedialog.asksaveasfilename(
            title="Save PDFRecon Case As",
            defaultextension=".prc",
            filetypes=[("PDFRecon Case Files", "*.prc"), ("All files", "*.*")]
        )
        if not filepath:
            return
        
        try:
            self._write_case_to_file(filepath)
            logging.info(f"Successfully saved case to: {filepath}")

            # Update state and GUI to reflect the save
            self.current_case_filepath = filepath
            self.case_is_dirty = False
            self.dirty_notes.clear()
            self._apply_filter()

        except Exception as e:
            logging.error(f"Failed to save case file '{filepath}': {e}")
            messagebox.showerror(self._("case_save_error_title"), self._("case_save_error_msg").format(e=e))
        
    def _export_reader(self):
        """
        Exports a fully self-contained Reader package, preserving the relative
        folder structure of the original scan inside the 'Evidence' folder.
        """
        if not self.all_scan_data or not self.last_scan_folder:
            messagebox.showwarning(self._("case_nothing_to_save_title"), self._("case_nothing_to_save_msg"))
            return
        
        base_path_str = filedialog.askdirectory(title=self._("export_reader_title"))
        if not base_path_str:
            return
        
        failed_file = ""
        dest_folder = Path(base_path_str) / "Export"

        try:
            if dest_folder.exists():
                shutil.rmtree(dest_folder)
            dest_folder.mkdir(exist_ok=True)

            evidence_folder = dest_folder / "Evidence"
            evidence_folder.mkdir(exist_ok=True)

            data_for_export = copy.deepcopy(self.all_scan_data)
            new_exif, new_timeline, new_hashes, path_map = {}, {}, {}, {}

            scan_base_path = self._resolve_case_path(self.last_scan_folder)

            for item in data_for_export:
                original_path_str = str(item['path'])
                original_abs_path = self._resolve_case_path(original_path_str)
                
                if not original_abs_path or not original_abs_path.exists():
                    logging.warning(f"Skipping missing file for export: {original_abs_path}")
                    continue

                try:
                    relative_sub_path = original_abs_path.relative_to(scan_base_path)
                except ValueError:
                    relative_sub_path = Path(original_abs_path.name)
                
                dest_file_path = evidence_folder / relative_sub_path
                dest_file_path.parent.mkdir(parents=True, exist_ok=True)

                failed_file = original_abs_path.name
                shutil.copy2(original_abs_path, dest_file_path)
                os.chmod(dest_file_path, stat.S_IREAD | stat.S_IRGRP | stat.S_IROTH)

                new_relative_path_str = str(Path("Evidence") / relative_sub_path)
                
                new_hashes[new_relative_path_str] = self._hash_file(dest_file_path)
                if original_path_str in self.exif_outputs:
                    new_exif[new_relative_path_str] = self.exif_outputs[original_path_str]
                if original_path_str in self.timeline_data:
                    new_timeline[new_relative_path_str] = self.timeline_data[original_path_str]

                path_map[original_path_str] = new_relative_path_str
                item['path'] = new_relative_path_str

            for item in data_for_export:
                if item.get('is_revision'):
                    original_parent_path = str(item.get('original_path'))
                    if original_parent_path in path_map:
                        item['original_path'] = path_map[original_parent_path]

            # --- Remap annotation keys to the new relative paths ---
            new_annotations = {}
            for original_path, note in self.file_annotations.items():
                if original_path in path_map:
                    new_annotations[path_map[original_path]] = note

            case_filename = f"case_{datetime.now().strftime('%Y%m%d_%H%M%S')}.prc"
            dest_case_file = dest_folder / case_filename
            failed_file = case_filename
            
            case_payload = {
                'app_version': self.app_version,
                'scan_folder': self.last_scan_folder,
                'all_scan_data': data_for_export,
                'file_annotations': new_annotations, # Use the remapped dictionary
                'exif_outputs': new_exif,
                'timeline_data': new_timeline,
                'path_to_id': self.path_to_id,
                'revision_counter': self.revision_counter,
                'evidence_hashes': new_hashes,
            }
            with open(dest_case_file, 'wb') as f:
                pickle.dump(case_payload, f)

            source_exe = Path(sys.executable)
            reader_exe_name = f"{source_exe.stem}_Reader{source_exe.suffix}"
            dest_exe = dest_folder / reader_exe_name
            failed_file = reader_exe_name
            shutil.copy2(source_exe, dest_exe)

            dependencies = ["license.txt", "config.ini", "icon.ico"]
            for dep_name in dependencies:
                source_dep = self._resolve_path(dep_name, base_is_parent=True)
                if source_dep.exists():
                    failed_file = dep_name
                    shutil.copy2(source_dep, dest_folder / dep_name)

            logging.info(f"Reader exported successfully to {dest_folder}")
            if messagebox.askyesno(self._("export_reader_success_title"), self._("export_reader_success_msg")):
                webbrowser.open(dest_folder)

        except Exception as e:
            logging.error(f"Failed to export Reader during operation on '{failed_file}': {e}")
            messagebox.showerror(
                self._("export_reader_error_title"),
                self._("export_reader_error_specific_msg").format(filename=failed_file, e=e)
            )
            
    def _find_pdf_files_generator(self, folder):
        """A generator that 'yields' PDF files as soon as they are found."""
        for base, _, files in os.walk(folder):
            for fn in files:
                if fn.lower().endswith(".pdf"):
                    yield Path(base) / fn

    def validate_pdf_file(self, filepath):
        """
        Validates a PDF file based on size, header, and encryption.
        Returns True if valid, otherwise raises an appropriate exception.
        """
        try:
            # Check file size
            if filepath.stat().st_size > PDFReconConfig.MAX_FILE_SIZE:
                raise PDFTooLargeError(f"File exceeds {PDFReconConfig.MAX_FILE_SIZE // (1024*1024)}MB size limit.")

            # Check PDF header
            with open(filepath, 'rb') as f:
                if f.read(5) != b'%PDF-':
                    raise PDFCorruptionError("Invalid PDF header. Not a PDF file.")
            
            # Check for encryption
            with fitz.open(filepath) as doc:
                if doc.is_encrypted:
                    # Try authenticating with an empty password
                    if not doc.authenticate(""):
                         raise PDFEncryptedError("File is encrypted and cannot be processed.")
            
            return True
        except PDFProcessingError:
            # Re-raise our custom exceptions
            raise
        except Exception as e:
            # Wrap other exceptions in our custom error type
            raise PDFCorruptionError(f"Could not validate file: {e}")

    def _process_large_file_streaming(self, filepath):
        """(Placeholder) Processes a large PDF file using streaming to save memory."""
        logging.info(f"Streaming processing is not yet implemented. Processing {filepath.name} normally.")
        pass
    def extract_additional_xmp_ids(self, txt: str) -> dict:
        """
        Harvest XMP IDs beyond basic xmpMM:{Original,Document,Instance}:
        - stRef:documentID / stRef:instanceID   (both attribute AND element forms)
        - xmpMM:DerivedFrom   (stRef:* and OriginalDocumentID inside)
        - xmpMM:Ingredients   (stRef:* inside)
        - xmpMM:History       (stRef:* and any InstanceID-like values)
        - photoshop:DocumentAncestors (Bag of GUIDs/uuids)
        Returns a dict of sets with normalized uppercase IDs.
        """
        def _norm(val):
            """Normalizes a UUID/GUID value for consistent comparison."""
            if val is None:
                return None
            if isinstance(val, (bytes, bytearray)):
                val = val.decode("utf-8", "ignore")
            s = str(val).strip()
            # strip known prefixes & wrappers
            s = re.sub(r"^urn:uuid:", "", s, flags=re.I)
            s = re.sub(r"^(uuid:|xmp\.iid:|xmp\.did:)", "", s, flags=re.I)
            s = s.strip("<>").strip()
            return s.upper() if s else None

        out = {
            "stref_doc_ids": set(),
            "stref_inst_ids": set(),
            "derived_doc_ids": set(),
            "derived_inst_ids": set(),
            "derived_orig_ids": set(),
            "ingredient_doc_ids": set(),
            "ingredient_inst_ids": set(),
            "history_inst_ids": set(),
            "history_doc_ids": set(),
            "ps_doc_ancestors": set(),
        }

        # ---------------- stRef anywhere (attribute form) ----------------
        for m in re.finditer(r'stRef:documentID="([^"]+)"', txt, re.I):
            v = _norm(m.group(1));  out["stref_doc_ids"].add(v) if v else None
        for m in re.finditer(r'stRef:instanceID="([^"]+)"', txt, re.I):
            v = _norm(m.group(1));  out["stref_inst_ids"].add(v) if v else None

        # ---------------- stRef anywhere (element form) ----------------
        for m in re.finditer(r"<stRef:documentID>([^<]+)</stRef:documentID>", txt, re.I):
            v = _norm(m.group(1));  out["history_doc_ids"].add(v) if v else None
        for m in re.finditer(r"<stRef:instanceID>([^<]+)</stRef:instanceID>", txt, re.I):
            v = _norm(m.group(1));  out["history_inst_ids"].add(v) if v else None

        # ---------------- DerivedFrom block ----------------
        df = re.search(r"<xmpMM:DerivedFrom\b[^>]*>(.*?)</xmpMM:DerivedFrom>", txt, re.I | re.S)
        if df:
            blk = df.group(1)
            # stRef:* within DerivedFrom
            for m in re.finditer(r'stRef:documentID="([^"]+)"', blk, re.I):
                v = _norm(m.group(1)); out["derived_doc_ids"].add(v) if v else None
            for m in re.finditer(r'stRef:instanceID="([^"]+)"', blk, re.I):
                v = _norm(m.group(1)); out["derived_inst_ids"].add(v) if v else None
            for m in re.finditer(r"<stRef:documentID>([^<]+)</stRef:documentID>", blk, re.I):
                v = _norm(m.group(1)); out["derived_doc_ids"].add(v) if v else None
            for m in re.finditer(r"<stRef:instanceID>([^<]+)</stRef:instanceID>", blk, re.I):
                v = _norm(m.group(1)); out["derived_inst_ids"].add(v) if v else None
            # OriginalDocumentID sometimes appears explicitly inside DerivedFrom
            for m in re.finditer(r"(?:xmpMM:|)OriginalDocumentID(?:>|=\")([^<\">]+)", blk, re.I):
                v = _norm(m.group(1)); out["derived_orig_ids"].add(v) if v else None

        # ---------------- Ingredients block ----------------
        ing = re.search(r"<xmpMM:Ingredients\b[^>]*>(.*?)</xmpMM:Ingredients>", txt, re.I | re.S)
        if ing:
            blk = ing.group(1)
            for m in re.finditer(r'stRef:documentID="([^"]+)"', blk, re.I):
                v = _norm(m.group(1)); out["ingredient_doc_ids"].add(v) if v else None
            for m in re.finditer(r'stRef:instanceID="([^"]+)"', blk, re.I):
                v = _norm(m.group(1)); out["ingredient_inst_ids"].add(v) if v else None
            for m in re.finditer(r"<stRef:documentID>([^<]+)</stRef:documentID>", blk, re.I):
                v = _norm(m.group(1)); out["ingredient_doc_ids"].add(v) if v else None
            for m in re.finditer(r"<stRef:instanceID>([^<]+)</stRef:instanceID>", blk, re.I):
                v = _norm(m.group(1)); out["ingredient_inst_ids"].add(v) if v else None

        # ---------------- History block ----------------
        hist = re.search(r"<xmpMM:History\b[^>]*>(.*?)</xmpMM:History>", txt, re.I | re.S)
        if hist:
            blk = hist.group(1)
            # attribute-style summary sometimes exists in certain producers
            for m in re.finditer(r'(?:InstanceID|stRef:instanceID)="([^"]+)"', blk, re.I):
                v = _norm(m.group(1)); out["history_inst_ids"].add(v) if v else None
            for m in re.finditer(r'(?:DocumentID|stRef:documentID)="([^"]+)"', blk, re.I):
                v = _norm(m.group(1)); out["history_doc_ids"].add(v) if v else None
            # element-style
            for m in re.finditer(r"<stRef:instanceID>([^<]+)</stRef:instanceID>", blk, re.I):
                v = _norm(m.group(1)); out["history_inst_ids"].add(v) if v else None
            for m in re.finditer(r"<stRef:documentID>([^<]+)</stRef:documentID>", blk, re.I):
                v = _norm(m.group(1)); out["history_doc_ids"].add(v) if v else None
            # catch any xmp.iid/xmp.did that appear as plain text within History
            for m in re.finditer(r"(uuid:[0-9a-f\-]+|xmp\.iid:[^,<>} \]]+|xmp\.did:[^,<>} \]]+)", blk, re.I):
                v = _norm(m.group(1)); out["history_inst_ids"].add(v) if v else None

        # ---------------- Photoshop DocumentAncestors ----------------
        ps = re.search(r"<photoshop:DocumentAncestors\b[^>]*>(.*?)</photoshop:DocumentAncestors>", txt, re.I | re.S)
        if ps:
            for m in re.finditer(r"<rdf:li[^>]*>([^<]+)</rdf:li>", ps.group(1), re.I):
                v = _norm(m.group(1)); out["ps_doc_ancestors"].add(v) if v else None

        # Deduplicate and purge empty values
        for k in out:
            out[k] = {v for v in out[k] if v}

        return out
    
    def _scan_worker_parallel(self, folder, q):
        """
        Finds PDF files and processes them in parallel using a ThreadPoolExecutor.
        Results are sent back to the main thread via a queue.
        """
        try:
            q.put(("scan_status", self._("preparing_analysis")))
            
            pdf_files = list(self._find_pdf_files_generator(folder))
            if not pdf_files:
                q.put(("finished", None))
                return

            q.put(("progress_mode_determinate", len(pdf_files)))
            files_processed = 0

            with ThreadPoolExecutor(max_workers=PDFReconConfig.MAX_WORKER_THREADS) as executor:
                # Pass the scan root 'folder' to each file processing job
                future_to_path = {executor.submit(self._process_single_file, fp, folder): fp for fp in pdf_files}
                for future in as_completed(future_to_path):
                    path = future_to_path[future]
                    files_processed += 1
                    try:
                        results = future.result()
                        for result_data in results:
                            q.put(("file_row", result_data))
                    except Exception as e:
                        logging.error(f"Unexpected error from thread pool for file {path.name}: {e}")
                        q.put(("file_row", {"path": path, "status": "error", "error_type": "unknown_error", "error_message": str(e)}))
                    
                    elapsed_time = time.time() - self.scan_start_time
                    fps = files_processed / elapsed_time if elapsed_time > 0 else 0
                    eta_seconds = (len(pdf_files) - files_processed) / fps if fps > 0 else 0
                    q.put(("detailed_progress", {"file": path.name, "fps": fps, "eta": time.strftime('%M:%S', time.gmtime(eta_seconds))}))

        except Exception as e:
            logging.error(f"Error in scan worker: {e}")
            q.put(("error", f"A critical error occurred: {e}"))
        finally:
            q.put(("finished", None))

    def _process_queue(self):
        """
        Processes messages from the worker thread's queue to update the GUI.
        This function runs periodically via `root.after()`.
        """
        try:
            # Process all available messages in the queue
            while True:
                msg_type, data = self.scan_queue.get_nowait()
                
                if msg_type == "progress_mode_determinate":
                    self.progressbar.config(mode='determinate', maximum=data if data > 0 else 1, value=0)
                elif msg_type == "detailed_progress":
                    self.progressbar['value'] += 1
                    self.status_var.set(self._("scan_progress_eta").format(**data))
                elif msg_type == "scan_status": 
                    self.status_var.set(data)
                elif msg_type == "file_row":
                    # Store all scan data, including revisions and errors
                    self.all_scan_data.append(data)
                    # Store EXIF and timeline data in separate dicts for quick lookup
                    if not data.get("is_revision"):
                        self.exif_outputs[str(data["path"])] = data.get("exif")
                        self.timeline_data[str(data["path"])] = data.get("timeline")
                    else: # For revisions
                        self.exif_outputs[str(data["path"])] = data.get("exif")
                        self.timeline_data[str(data["path"])] = data.get("timeline")
                        self.revision_counter += 1

                elif msg_type == "error": 
                    logging.warning(data)
                    messagebox.showerror("Critical Error", data)
                elif msg_type == "finished":
                    self._finalize_scan()
                    return # Stop polling the queue
        except queue.Empty:
            # If the queue is empty, do nothing
            pass
        # Schedule the next check
        self.root.after(100, self._process_queue)
    
    def _finalize_scan(self):
        """Performs final actions after a scan is complete."""
        # Repopulate the tree with the final data
        self._apply_filter()
        
        # --- Update GUI to 'finished' state ---
        self.scan_button.config(state="normal")
        self.export_menubutton.config(state="normal")

        # Calculate hashes for the evidence files for integrity verification
        self.evidence_hashes = self._calculate_hashes(self.all_scan_data)
        if self.evidence_hashes:
             self.file_menu.entryconfig(self._("menu_verify_integrity"), state="normal")

        # Enable saving and exporting options in the full program
        if not self.is_reader_mode:
            self.file_menu.entryconfig(self._("menu_save_case"), state="normal")
            if getattr(sys, 'frozen', False):
                 self.file_menu.entryconfig(self._("menu_export_reader"), state="normal")
        
        # Ensure the progress bar is full
        if self.progressbar['value'] < self.progressbar['maximum']:
             self.progressbar['value'] = self.progressbar['maximum']
        
        # Wait for any lingering background file copy operations to finish
        # and then update the status bar with the final summary.
        self._finalize_copy_operations()
        
        
    def _apply_filter(self, *args):
            """Filters the displayed results based on the search term."""
            search_term = self.filter_var.get().lower()
            
            items_to_show = []
            if not search_term:
                items_to_show = self.all_scan_data
            else:
                for data in self.all_scan_data:
                    searchable_items = []

                    path_str = str(data.get('path', ''))
                    searchable_items.append(path_str)
                    searchable_items.append(data.get('md5', ''))

                    # Add data from File Created and File Modified columns
                    if not data.get('is_revision'):
                        try:
                            # Resolve path to be absolute before calling .stat()
                            resolved_path = self._resolve_case_path(data['path'])
                            if resolved_path and resolved_path.exists():
                                stat = resolved_path.stat()
                                searchable_items.append(datetime.fromtimestamp(stat.st_ctime).strftime("%d-%m-%Y %H:%M:%S"))
                                searchable_items.append(datetime.fromtimestamp(stat.st_mtime).strftime("%d-%m-%Y %H:%M:%S"))
                        except (FileNotFoundError, KeyError, AttributeError):
                            pass # Safely ignore if file not found or path is not valid

                    # Add data from the 'Altered' column
                    is_rev = data.get("is_revision", False)
                    if data.get("status") == "error":
                        error_type_key = data.get("error_type", "unknown_error")
                        searchable_items.append(self._(error_type_key))
                    elif is_rev:
                        if data.get("is_identical"):
                             searchable_items.append(self._("status_identical"))
                        searchable_items.append(self._("revision_of").split("{")[0])
                    else: # For original files
                        flag = self.get_flag(data.get("indicator_keys", {}), False)
                        searchable_items.append(flag)

                    # Add the full EXIFTool output
                    exif_output = self.exif_outputs.get(path_str, '')
                    if exif_output:
                        searchable_items.append(exif_output)

                    # Add the detailed indicator text from the "Signs of Alteration" column
                    indicator_dict = data.get('indicator_keys', {})
                    if indicator_dict:
                        indicator_details = [self._format_indicator_details(key, details) for key, details in indicator_dict.items()]
                        searchable_items.extend(indicator_details)
                    elif not is_rev:
                        searchable_items.append(self._("status_no"))
                    
                    full_searchable_text = " ".join(searchable_items).lower()
                    if search_term in full_searchable_text:
                        items_to_show.append(data)
            
            self._populate_tree_from_data(items_to_show)  
            
    def _populate_tree_from_data(self, data_list):
        """
        Repopulates the treeview from a (filtered) list of rows in a flat structure.
        """
        self.tree.delete(*self.tree.get_children())
        self.report_data.clear()

        # Pre-calculate the display IDs for parent files, used for "Revision of #X" text
        parent_display_ids = {}
        parent_counter = 0
        for d in self.all_scan_data: # Use all_scan_data for stable IDs
            if not d.get("is_revision") and d.get("status") != "error":
                parent_counter += 1
                parent_display_ids[str(d["path"])] = parent_counter

        # Populate the tree in a single pass using the filtered data_list
        for i, d in enumerate(data_list):
            path_obj = Path(d["path"])
            path_str = str(d["path"])
            is_rev = d.get("is_revision", False)
            indicator_keys = d.get("indicator_keys", {})

            # --- For each row, check for a note ---
            # This logic runs for every single file being displayed.
            # It checks if its specific path exists in the notes dictionaries.
            note_indicator = ""
            if path_str in self.dirty_notes:
                note_indicator = "📝*"
            elif path_str in self.file_annotations:
                note_indicator = "📝"

            exif_display = "✔" if d.get("exif") else ""

            if is_rev:
                parent_display_id = parent_display_ids.get(str(d.get("original_path")))
                display_id = parent_display_id if parent_display_id else i + 1
                flag = self._("status_identical").format(pages=PDFReconConfig.VISUAL_DIFF_PAGE_LIMIT) if d.get("is_identical") else self.get_flag({}, True, parent_display_id)
                tag = "gray_row" if d.get("is_identical") else "blue_row"
                revisions_display, created_time, modified_time, indicators_display = "", "", "", ""
            else: # Is a parent file
                display_id = parent_display_ids.get(path_str, i + 1)
                flag = self.get_flag(indicator_keys, False)
                tag = self.tree_tags.get(flag, "")
                revisions_count = indicator_keys.get("HasRevisions", {}).get("count", 0)
                revisions_display = str(revisions_count) if revisions_count > 0 else ""
                indicators_display = "✔" if indicator_keys else ""
                try:
                    full_path = self._resolve_case_path(path_obj)
                    st = full_path.stat()
                    created_time = datetime.fromtimestamp(st.st_ctime).strftime("%d-%m-%Y %H:%M:%S")
                    modified_time = datetime.fromtimestamp(st.st_mtime).strftime("%d-%m-%Y %H:%M:%S")
                except Exception:
                    created_time, modified_time = "", ""

            row_values = [
                display_id, path_obj.name, flag, revisions_display, path_str,
                d.get("md5", ""), created_time, modified_time,
                exif_display, indicators_display, note_indicator
            ]
            
            self.tree.insert("", "end", values=row_values, tags=(tag,))
            self.report_data.append(row_values)
            
    def on_select_item(self, event):
        """Updates the detail view and the live inspector window when an item is selected."""
        selected_items = self.tree.selection()
        if not selected_items:
            self.detail_text.config(state="normal")
            self.detail_text.delete("1.0", tk.END)
            self.detail_text.config(state="disabled")
            return
        
        item_id = selected_items[0]
        values = self.tree.item(item_id, "values")
        path_str = values[4]
        
        self.detail_text.config(state="normal")
        self.detail_text.delete("1.0", tk.END)
        
        original_data = next((d for d in self.all_scan_data if str(d.get('path')) == path_str), None)

        for i, val in enumerate(values):
            col_name = self.tree.heading(self.columns[i], "text")
            self.detail_text.insert(tk.END, f"{col_name}: ", ("bold",))
            
            if col_name == self._("col_path"):
                self.detail_text.insert(tk.END, val + "\n", ("link",))
            elif col_name == self._("col_indicators") and original_data and original_data.get("indicator_keys"):
                indicator_details = [self._format_indicator_details(key, details) for key, details in original_data["indicator_keys"].items()]
                full_indicators_str = "\n  • " + "\n  • ".join(indicator_details)
                self.detail_text.insert(tk.END, full_indicators_str + "\n")
            else:
                self.detail_text.insert(tk.END, val + "\n")
                
        note = self.file_annotations.get(path_str)
        if note:
            self.detail_text.insert(tk.END, "\n" + "-"*40 + "\n")
            self.detail_text.insert(tk.END, f"{self._('note_label')}\n", ("bold",))
            self.detail_text.insert(tk.END, note)

        self.detail_text.config(state="disabled")

        # --- NY KODE: Opdater Inspector-vinduet, hvis det er åbent ---
        if self.inspector_window and self.inspector_window.winfo_viewable():
            self.show_inspector_popup()            

    def _open_path_from_detail(self, event):
        """Opens the folder when a path link is clicked in the detail view."""
        index = self.detail_text.index(f"@{event.x},{event.y}")
        # Find which link tag was clicked
        tag_indices = self.detail_text.tag_ranges("link")
        for start, end in zip(tag_indices[0::2], tag_indices[1::2]):
            if self.detail_text.compare(start, "<=", index) and self.detail_text.compare(index, "<", end):
                path_str = self.detail_text.get(start, end).strip()
                try:
                    webbrowser.open(os.path.dirname(path_str))
                except Exception as e:
                    messagebox.showerror(self._("open_folder_error_title"), f"Could not open the folder: {e}")
                break

    def extract_revisions(self, raw, original_path):
        """
        Extracts previous versions (revisions) of a PDF from its raw byte content
        by looking for '%%EOF' markers. It prepares potential paths but does not write files.
        """
        revisions = []
        offsets = []
        pos = len(raw)
        # Find all '%%EOF' markers from the end of the file backwards
        while (pos := raw.rfind(b"%%EOF", 0, pos)) != -1: offsets.append(pos)
        
        # Filter out invalid or unlikely offsets
        valid_offsets = [o for o in sorted(offsets) if 1000 <= o <= len(raw) - 500]
        if valid_offsets:
            # Define the subdirectory for potential revisions
            altered_dir = original_path.parent / "Altered_files"
            
            for i, offset in enumerate(valid_offsets, start=1):
                rev_bytes = raw[:offset + 5] # The revision is the content from the start to the EOF marker
                rev_filename = f"{original_path.stem}_rev{i}_@{offset}.pdf"
                rev_path = altered_dir / rev_filename
                # Package the data for later validation without writing the file.
                revisions.append((rev_path, original_path.name, rev_bytes))
                
        return revisions

    def exiftool_output(self, path, detailed=False):
        """Runs ExifTool safely with a timeout and improved error handling."""
        exe_path = self._resolve_path("exiftool.exe", base_is_parent=True)
        if not exe_path.is_file(): return self._("exif_err_notfound")
        
        try:
            file_content = path.read_bytes()
            # Suppress console window on Windows
            startupinfo = None
            if sys.platform == "win32":
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            
            # Build the command-line arguments for ExifTool
            command = [str(exe_path)]
            if detailed: command.extend(["-a", "-s", "-G1", "-struct"])
            else: command.extend(["-a"])
            command.append("-") # Read from stdin

            # Run the process
            process = subprocess.run(
                command,
                input=file_content,
                capture_output=True,
                check=False,
                startupinfo=startupinfo,
                timeout=PDFReconConfig.EXIFTOOL_TIMEOUT
            )
            
            # Handle non-zero exit codes or stderr output
            if process.returncode != 0 or process.stderr:
                error_message = process.stderr.decode('latin-1', 'ignore').strip()
                if not process.stdout.strip(): return f"{self._('exif_err_prefix')}\n{error_message}"
                logging.warning(f"ExifTool stderr for {path.name}: {error_message}")

            # Decode the output, trying UTF-8 first, then latin-1 as a fallback
            try: raw_output = process.stdout.decode('utf-8').strip()
            except UnicodeDecodeError: raw_output = process.stdout.decode('latin-1', 'ignore').strip()

            # Remove empty lines from the output
            return "\n".join([line for line in raw_output.splitlines() if line.strip()])

        except subprocess.TimeoutExpired:
            logging.error(f"ExifTool timed out for file {path.name}")
            return self._("exif_err_prefix") + f"\nTimeout after {PDFReconConfig.EXIFTOOL_TIMEOUT} seconds."
        except Exception as e:
            logging.error(f"Error running exiftool for file {path}: {e}")
            return self._("exif_err_run").format(e=e)

    def _get_filesystem_times(self, filepath):
            """Helper function to get created/modified timestamps from the file system."""
            events = []
            try:
                stat = filepath.stat()
                mtime = datetime.fromtimestamp(stat.st_mtime)
                events.append((mtime, f"File System: {self._('col_modified')}"))
                ctime = datetime.fromtimestamp(stat.st_ctime)
                events.append((ctime, f"File System: {self._('col_created')}"))
            except FileNotFoundError:
                pass
            return events

    def _parse_exif_data(self, exiftool_output: str):
        """
        Parses EXIFTool output into a structured dictionary for reuse.
        """
        data = {
            "producer_pdf": "", "producer_xmppdf": "", "softwareagent": "",
            "application": "", "software": "", "creatortool": "", "xmptoolkit": "",
            "create_dt": None, "modify_dt": None, "history_events": [], "all_dates": []
        }
        lines = exiftool_output.splitlines()

        # --- Regex Patterns ---
        kv_re = re.compile(r'^\[(?P<group>[^\]]+)\]\s*(?P<tag>[\w\-/ ]+?)\s*:\s*(?P<value>.+)$')
        date_re = re.compile(r'^(?P<date>\d{4}:\d{2}:\d{2}\s\d{2}:\d{2}:\d{2}(?:[+\-]\d{2}:\d{2}|Z)?)')
        history_full_pattern = re.compile(r"\[XMP-xmpMM\]\s+History\s+:\s+(.*)")

        def looks_like_software(s: str) -> bool:
            return bool(s and self.software_tokens.search(s))

        # --- First Pass: Collect Key-Value Pairs for Tools ---
        for ln in lines:
            m = kv_re.match(ln)
            if not m: continue
            
            group = m.group("group").strip().lower()
            tag = m.group("tag").strip().lower().replace(" ", "")
            val = m.group("value").strip()

            if tag == "producer":
                if group == "pdf" and not data["producer_pdf"]: data["producer_pdf"] = val
                elif group in ("xmp-pdf", "xmp_pdf") and not data["producer_xmppdf"]: data["producer_xmppdf"] = val
            elif tag == "softwareagent" and not data["softwareagent"]: data["softwareagent"] = val
            elif tag == "application" and not data["application"]: data["application"] = val
            elif tag == "software" and not data["software"]: data["software"] = val
            elif tag == "creatortool" and not data["creatortool"] and looks_like_software(val):
                data["creatortool"] = val
            elif tag == "xmptoolkit" and not data["xmptoolkit"]: data["xmptoolkit"] = val
        
        # Fallback for producer fields
        if not data["producer_pdf"] and data["producer_xmppdf"]: data["producer_pdf"] = data["producer_xmppdf"]
        if not data["producer_xmppdf"] and data["producer_pdf"]: data["producer_xmppdf"] = data["producer_pdf"]

        # --- Second Pass: Collect All Dates and History Events ---
        for ln in lines:
            # History events
            hist_match = history_full_pattern.match(ln)
            if hist_match:
                history_str = hist_match.group(1)
                event_blocks = re.findall(r'\{([^}]+)\}', history_str)
                for block in event_blocks:
                    details = {k.strip(): v.strip() for k, v in (pair.split('=', 1) for pair in block.split(',') if '=' in pair)}
                    if 'When' in details:
                        try:
                            dt_obj = datetime.strptime(details['When'].replace("Z", "+00:00").split('+')[0].split('.')[0], "%Y:%m:%d %H:%M:%S")
                            data["history_events"].append((dt_obj, details))
                        except (ValueError, IndexError):
                            pass
                continue

            # Generic date lines
            kv_match = kv_re.match(ln)
            if not kv_match: continue

            val_str = kv_match.group("value").strip()
            date_match = date_re.match(val_str)
            if date_match:
                try:
                    date_str = date_match.group("date")
                    base_date = date_str.replace("Z", "+00:00").split('+')[0].split('-')[0]
                    dt = datetime.strptime(base_date, "%Y:%m:%d %H:%M:%S")
                    
                    tag = kv_match.group("tag").strip().lower().replace(" ", "")
                    group = kv_match.group("group").strip()
                    data["all_dates"].append({"dt": dt, "tag": tag, "group": group, "full_str": date_str})

                except ValueError:
                    continue
        
        # Find first create date and latest modify date
        for d in data["all_dates"]:
            if d["tag"] in {"createdate", "creationdate"}:
                if data["create_dt"] is None or d["dt"] < data["create_dt"]:
                    data["create_dt"] = d["dt"]
            elif d["tag"] in {"modifydate", "metadatadate"}:
                if data["modify_dt"] is None or d["dt"] > data["modify_dt"]:
                    data["modify_dt"] = d["dt"]
        
        return data

    def _detect_tool_change_from_exif(self, exiftool_output: str):
        """
        Determines if the primary tool changed between creation and last modification.
        This function is now a lightweight wrapper around _parse_exif_data.
        """
        data = self._parse_exif_data(exiftool_output)
        
        create_tool = data["producer_pdf"] or data["producer_xmppdf"] or data["application"] or data["software"] or data["creatortool"] or ""
        modify_tool = data["softwareagent"] or data["producer_pdf"] or data["producer_xmppdf"] or data["application"] or data["software"] or data["creatortool"] or ""
        
        create_engine = modify_engine = ""
        if data["xmptoolkit"]:
            if data["create_dt"]: create_engine = data["xmptoolkit"]
            if data["modify_dt"]: modify_engine = data["xmptoolkit"]

        changed_tool = bool(create_tool and modify_tool and create_tool.strip() != modify_tool.strip())
        changed_engine = bool(create_engine and modify_engine and create_engine.strip() != modify_engine.strip())
        
        reason = ""
        if changed_tool and changed_engine: reason = "mixed"
        elif changed_tool: reason = "producer" if (data["producer_pdf"] or data["producer_xmppdf"]) else "software"
        elif changed_engine: reason = "engine"

        return {
            "changed": changed_tool or changed_engine,
            "create_tool": create_tool, "modify_tool": modify_tool,
            "create_engine": create_engine, "modify_engine": modify_engine,
            "modify_dt": data["modify_dt"],
            "reason": reason
        }

    def _parse_exiftool_timeline(self, exiftool_output):
        """
        Generates a list of timeline events from parsed EXIF data.
        """
        events = []
        data = self._parse_exif_data(exiftool_output)

        create_tool = data["producer_pdf"] or data["producer_xmppdf"] or data["application"] or data["software"] or data["creatortool"] or ""
        modify_tool = data["softwareagent"] or create_tool # Fallback to create_tool if no specific modify tool

        # --- Add History Events ---
        for dt_obj, details in data["history_events"]:
            action = details.get('Action', 'N/A')
            agent = details.get('SoftwareAgent', '')
            changed = details.get('Changed', '')
            desc = [f"Action: {action}"]
            if agent: desc.append(f"Agent: {agent}")
            if changed: desc.append(f"Changed: {changed}")
            events.append((dt_obj, f"XMP History   - {' | '.join(desc)}"))

        # --- Add Generic Date Events ---
        def _ts_label(tag: str) -> str:
            t = tag.replace(" ", "").lower()
            return {"createdate": "Created", "creationdate": "Created", "modifydate": "Modified", "metadatadate": "Metadata"}.get(t, tag)

        for d in data["all_dates"]:
            label = self._(_ts_label(d["tag"]).lower())
            tool = create_tool if d["tag"] in {"createdate", "creationdate"} else modify_tool
            tool_part = f" | Tool: {tool}" if tool else ""
            events.append((d["dt"], f"ExifTool ({d['group']}) - {label}: {d['full_str']}{tool_part}"))
        
        # --- Add XMP Engine Information ---
        if data["xmptoolkit"]:
            anchor_dt = data["create_dt"] or (data["all_dates"][0]["dt"] if data["all_dates"] else datetime.now())
            label_engine = "XMP Engine" if self.language.get() == "en" else "XMP-motor"
            events.append((anchor_dt, f"{label_engine}: {data['xmptoolkit']}"))

        return events
        
    @staticmethod
    def _format_timedelta(delta):
        """Formats a timedelta object into a readable string (e.g., (+1d 2h 3m 4.56s))."""
        if not delta or delta.total_seconds() < 0.001:
            return ""

        s = delta.total_seconds()
        days, remainder = divmod(s, 86400)
        hours, remainder = divmod(remainder, 3600)
        minutes, seconds = divmod(remainder, 60)
        
        parts = []
        if days > 0: parts.append(f"{int(days)}d")
        if hours > 0: parts.append(f"{int(hours)}h")
        if minutes > 0: parts.append(f"{int(minutes)}m")
        if seconds > 0 or not parts: parts.append(f"{seconds:.2f}s")

        return f"(+{ ' '.join(parts) })"
            
    def _parse_raw_content_timeline(self, file_content_string):
        """Helper function to parse timestamps directly from the file's raw content."""
        events = []
        
        # Look for PDF-style dates: /CreationDate (D:20230101120000...)
        pdf_date_pattern = re.compile(r"\/([A-Z][a-zA-Z0-9_]+)\s*\(\s*D:(\d{14})")
        for match in pdf_date_pattern.finditer(file_content_string):
            label, date_str = match.groups()
            try:
                dt_obj = datetime.strptime(date_str, "%Y%m%d%H%M%S")
                display_line = f"Raw File: /{label}: {dt_obj.strftime('%Y-%m-%d %H:%M:%S')}"
                events.append((dt_obj, display_line))
            except ValueError:
                continue

        # Look for XMP-style dates: <xmp:CreateDate>2023-01-01T12:00:00Z</xmp:CreateDate>
        xmp_date_pattern = re.compile(r"<([a-zA-Z0-9:]+)[^>]*?>\s*(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}.*?)\s*<\/([a-zA-Z0-9:]+)>")
        for match in xmp_date_pattern.finditer(file_content_string):
            label, date_str, closing_label = match.groups()
            if label != closing_label: continue # Ensure tags match
            try:
                # Clean the date string of timezones and milliseconds
                clean_date_str = date_str.split('Z')[0].split('+')[0].split('.')[0].strip()
                dt_obj = datetime.fromisoformat(clean_date_str)
                display_line = f"Raw File: <{label}>: {date_str}"
                events.append((dt_obj, display_line))
            except (ValueError, IndexError):
                continue
        return events        
    def _clean_cell_value(self, value):
        """Fjerner tegn, der kan give XLSX/XML-fejl (BOM/mojibake/kontroltegn)."""
        import re
        if value is None:
            return ""
        s = str(value)
        # ulovlige XML-kontroltegn (tillader \t \n \r)
        s = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", s)
        # reelle BOM-tegn
        if s.startswith("\ufeff") or s.startswith("\ufffe"):
            s = s.lstrip("\ufeff\ufffe")
        # typisk mojibake for BOM ('þÿ'/'ÿþ')
        if s.startswith("þÿ") or s.startswith("ÿþ"):
            s = s[2:]
        # fjern NUL
        s = s.replace("\x00", "")
        return s

    def generate_comprehensive_timeline(self, filepath, raw_file_content, exiftool_output):
        """
        Combines events from all sources, separating them into timezone-aware and naive lists.
        """
        all_events = []

        # 1) Get File System, ExifTool, and Raw Content timestamps
        all_events.extend(self._get_filesystem_times(filepath))
        all_events.extend(self._parse_exiftool_timeline(exiftool_output))
        all_events.extend(self._parse_raw_content_timeline(raw_file_content))

        # 2) Add a special event if a tool change was detected
        try:
            info = self._detect_tool_change_from_exif(exiftool_output)
            if info.get("changed"):
                when = info.get("modify_dt")
                if not when and all_events:
                    # Find a datetime object to anchor the event, prioritizing naive ones if present
                    naive_dts = [e[0] for e in all_events if e[0].tzinfo is None]
                    when = max(naive_dts) if naive_dts else max(e[0] for e in all_events)
                if not when:
                    when = datetime.now()
                
                # Format the description of the tool change
                if self.language.get() == "da":
                    label = "Værktøj skiftet"
                    parts = [f"{info.get('create_tool','?')} → {info.get('modify_tool','?')}"]
                    if info.get("reason") == "engine":
                        parts.append(f"(XMP-motor: {info.get('create_engine','?')} → {info.get('modify_engine','?')})")
                    line = f"{label}: " + " ".join(parts)
                else:
                    label = "Tool changed"
                    parts = [f"{info.get('create_tool','?')} → {info.get('modify_tool','?')}"]
                    if info.get("reason") == "engine":
                        parts.append(f"(XMP engine: {info.get('create_engine','?')} → {info.get('modify_engine','?')})")
                    line = f"{label}: " + " ".join(parts)
                all_events.append((when, line))
        except Exception:
            pass

        # 3) Separate events into two lists based on timezone info
        aware_events = []
        naive_events = []
        for dt_obj, description in all_events:
            if dt_obj.tzinfo is not None:
                aware_events.append((dt_obj, description))
            else:
                naive_events.append((dt_obj, description))

        # 4) Sort each list independently
        aware_events.sort(key=lambda x: x[0])
        naive_events.sort(key=lambda x: x[0])

        return {"aware": aware_events, "naive": naive_events}
    
    def _populate_timeline_widget(self, text_widget, path_str):
        """Helper function to populate a text widget with formatted timeline data."""
        timeline_data = self.timeline_data.get(path_str)
        
        if not timeline_data or (not timeline_data.get("aware") and not timeline_data.get("naive")):
            text_widget.insert("1.0", self._("timeline_no_data"))
            return

        # Configure tags on the provided text_widget
        text_widget.tag_configure("date_header", font=("Courier New", 11, "bold", "underline"), spacing1=10, spacing3=5)
        text_widget.tag_configure("time", font=("Courier New", 10, "bold"))
        text_widget.tag_configure("delta", foreground="#0078D7")
        text_widget.tag_configure("section_header", font=("Courier New", 12, "bold"), spacing1=15, spacing3=10, justify='center')
        text_widget.tag_configure("source_fs", foreground="#008000")
        text_widget.tag_configure("source_exif", foreground="#555555")
        text_widget.tag_configure("source_raw", foreground="#800080")
        text_widget.tag_configure("source_xmp", foreground="#C00000")

        aware_events = timeline_data.get("aware", [])
        naive_events = timeline_data.get("naive", [])
        
        if aware_events:
            last_date = None
            last_dt_obj = None
            for dt_obj, description in aware_events:
                local_dt = dt_obj.astimezone()
                if local_dt.date() != last_date:
                    if last_date is not None: text_widget.insert("end", "\n")
                    text_widget.insert("end", f"--- {local_dt.strftime('%d-%m-%Y')} ---\n", "date_header")
                    last_date = local_dt.date()
                delta_str = ""
                if last_dt_obj:
                    delta = local_dt - last_dt_obj
                    delta_str = self._format_timedelta(delta)
                source_tag = "source_exif"
                if description.startswith("FS"): source_tag = "source_fs"
                time_str = local_dt.strftime('%H:%M:%S %z')
                text_widget.insert("end", f"{time_str:<15}", "time")
                text_widget.insert("end", f" | {description:<60}", source_tag)
                text_widget.insert("end", f" | {delta_str}\n", "delta")
                last_dt_obj = local_dt

        if naive_events:
            header_text = ("\n--- Tider uden tidszoneinformation ---\n" if self.language.get() == "da" 
                           else "\n--- Times without timezone information ---\n")
            text_widget.insert("end", header_text, "section_header")
            
            last_date = None
            last_dt_obj = None
            for dt_obj, description in naive_events:
                if dt_obj.date() != last_date:
                    if last_date is not None: text_widget.insert("end", "\n")
                    text_widget.insert("end", f"--- {dt_obj.strftime('%d-%m-%Y')} ---\n", "date_header")
                    last_date = dt_obj.date()
                delta_str = ""
                if last_dt_obj:
                    delta = dt_obj - last_dt_obj
                    delta_str = self._format_timedelta(delta)
                source_tag = "source_exif"
                if description.startswith("File System"): source_tag = "source_fs"
                elif description.startswith("Raw File"): source_tag = "source_raw"
                elif description.startswith("XMP History"): source_tag = "source_xmp"
                time_str = dt_obj.strftime('%H:%M:%S')
                text_widget.insert("end", f"{time_str:<15}", "time")
                text_widget.insert("end", f" | {description:<60}", source_tag)
                text_widget.insert("end", f" | {delta_str}\n", "delta")
                last_dt_obj = dt_obj
                
    @staticmethod
    def decompress_stream(b):
        """Attempts to decompress a PDF stream using common filters."""
        for fn in (zlib.decompress, lambda d: base64.a85decode(re.sub(rb"\s", b"", d), adobe=True), lambda d: binascii.unhexlify(re.sub(rb"\s|>", b"", d))):
            try: return fn(b).decode("latin1", "ignore")
            except Exception: pass
        return ""

    @staticmethod
    def extract_text(raw: bytes):
        """
        Extracts only what's needed for indicator hunting:
        - ~2 MB header/trailer
        - Small streams (skipping large image streams)
        - XMP xpacket (if present)
        """
        txt_segments = []

        # Cap: header/trailer/objects
        head_cap = raw[:2_000_000].decode("latin1", "ignore")
        txt_segments.append(head_cap)

        # Only process small streams (e.g., <= 256 KB) to avoid inflating large images
        for m in re.finditer(rb"stream\r?\n(.*?)\r?\nendstream", raw, re.S):
            body = m.group(1)
            if len(body) <= 256_000:
                try:
                    txt_segments.append(PDFReconApp.decompress_stream(body))
                except Exception:
                    try:
                        txt_segments.append(body.decode("latin1", "ignore"))
                    except Exception:
                        pass

        # XMP xpacket (full content)
        m = re.search(rb"<\?xpacket begin=.*?\?>(.*?)<\?xpacket end=[^>]*\?>", raw, re.S)
        if m:
            try:
                txt_segments.append(m.group(1).decode("utf-8", "ignore"))
            except Exception:
                txt_segments.append(m.group(1).decode("latin1", "ignore"))

        return "\n".join(txt_segments)

    def analyze_fonts(self, filepath, doc):
            """
            Analyzes fonts to detect multiple subsets of the same base font.
            Returns a dictionary of conflicting fonts, e.g.,
            {'Calibri': {'ABC+Calibri', 'DEF+Calibri-Bold'}}
            """
            font_subsets = {}
            # Iterate through each page to get the fonts used
            for page_num in range(len(doc)):
                fonts_on_page = doc.get_page_fonts(page_num)
                for font_info in fonts_on_page:
                    basefont_name = font_info[3]
                    if "+" in basefont_name:
                        try:
                            _, actual_base_font = basefont_name.split("+", 1)
                            normalized_base = actual_base_font.split('-')[0]
                            if normalized_base not in font_subsets:
                                font_subsets[normalized_base] = set()
                            font_subsets[normalized_base].add(basefont_name)
                        except ValueError:
                            continue
            
            # Filter for only those fonts that actually have multiple subsets
            conflicting_fonts = {base: subsets for base, subsets in font_subsets.items() if len(subsets) > 1}
            if conflicting_fonts:
                logging.info(f"Multiple font subsets found in {filepath.name}: {conflicting_fonts}")

            return conflicting_fonts
    def detect_indicators(self, filepath, txt: str, doc):
        """
        Searches for indicators, now with an integrated text-diff feature for TouchUp edits.
        """
        indicators = {}

        # --- High-Confidence Indicators ---
        if re.search(r"touchup_textedit", txt, re.I):
            found_text = self._extract_touchup_text(doc)
            details = {'found_text': found_text, 'text_diff': None}

            # Check for revisions to perform a text diff
            revisions = self.extract_revisions(doc.write(), filepath)
            if revisions:
                latest_rev_path, _, latest_rev_bytes = revisions[0] # Compare against the most recent revision
                logging.info(f"TouchUp found with revisions. Performing text diff for {filepath.name}.")
                
                original_text = self._get_text_for_comparison(filepath)
                revision_text = self._get_text_for_comparison(latest_rev_bytes)

                if original_text and revision_text:
                    diff = difflib.unified_diff(
                        revision_text.splitlines(keepends=True),
                        original_text.splitlines(keepends=True),
                        fromfile='Previous Version',
                        tofile='Final Version',
                    )
                    details['text_diff'] = list(diff)
            indicators['TouchUp_TextEdit'] = details

        # --- Metadata Indicators ---
        creators = set(re.findall(r"/Creator\s*\((.*?)\)", txt, re.I))
        if len(creators) > 1:
            indicators['MultipleCreators'] = {'count': len(creators), 'values': list(creators)}
        
        producers = set(re.findall(r"/Producer\s*\((.*?)\)", txt, re.I))
        if len(producers) > 1:
            indicators['MultipleProducers'] = {'count': len(producers), 'values': list(producers)}

        if re.search(r'<xmpMM:History>', txt, re.I | re.S):
            indicators['XMPHistory'] = {}

        # --- Structural and Content Indicators ---
        try:
            conflicting_fonts = self.analyze_fonts(filepath, doc)
            if conflicting_fonts:
                indicators['MultipleFontSubsets'] = {'fonts': conflicting_fonts}
        except Exception as e:
            logging.error(f"Error analyzing fonts for {filepath.name}: {e}")

        if (hasattr(doc, 'is_xfa') and doc.is_xfa) or "/XFA" in txt:
            indicators['HasXFAForm'] = {}

        if re.search(r"/Type\s*/Sig\b", txt):
            indicators['HasDigitalSignature'] = {}

        # --- Incremental Update Indicators ---
        startxref_count = txt.lower().count("startxref")
        if startxref_count > 1:
            indicators['MultipleStartxref'] = {'count': startxref_count}
        
        prevs = re.findall(r"/Prev\s+\d+", txt)
        if prevs:
            indicators['IncrementalUpdates'] = {'count': len(prevs) + 1}
        
        if re.search(r"/Linearized\s+\d+", txt):
            indicators['Linearized'] = {}
            if startxref_count > 1 or prevs:
                indicators['LinearizedUpdated'] = {}

        # --- Feature Indicators ---
        if re.search(r"/Redact\b", txt, re.I): indicators['HasRedactions'] = {}
        if re.search(r"/Annots\b", txt, re.I): indicators['HasAnnotations'] = {}
        if re.search(r"/PieceInfo\b", txt, re.I): indicators['HasPieceInfo'] = {}
        if re.search(r"/AcroForm\b", txt, re.I):
            indicators['HasAcroForm'] = {}
            if re.search(r"/NeedAppearances\s+true\b", txt, re.I):
                indicators['AcroFormNeedAppearances'] = {}

        gen_gt_zero_matches = [m for m in re.finditer(r"\b(\d+)\s+(\d+)\s+obj\b", txt) if int(m.group(2)) > 0]
        if gen_gt_zero_matches:
            indicators['ObjGenGtZero'] = {'count': len(gen_gt_zero_matches)}

        # --- ID Comparison ---
        def _norm_uuid(x):
            if x is None: return None
            s = str(x).strip().upper()
            return re.sub(r"^(URN:UUID:|UUID:|XMP\.IID:|XMP\.DID:)", "", s).strip("<>")

        xmp_orig = _norm_uuid(re.search(r"xmpMM:OriginalDocumentID(?:>|=\")([^<\"]+)", txt, re.I).group(1) if re.search(r"xmpMM:OriginalDocumentID", txt, re.I) else None)
        xmp_doc = _norm_uuid(re.search(r"xmpMM:DocumentID(?:>|=\")([^<\"]+)", txt, re.I).group(1) if re.search(r"xmpMM:DocumentID", txt, re.I) else None)
        
        if xmp_orig and xmp_doc and xmp_doc != xmp_orig:
            indicators['XMPIDChange'] = {'from': xmp_orig, 'to': xmp_doc}

        trailer_match = re.search(r"/ID\s*\[\s*<\s*([0-9A-Fa-f]+)\s*>\s*<\s*([0-9A-Fa-f]+)\s*>\s*\]", txt)
        if trailer_match:
            trailer_orig, trailer_curr = _norm_uuid(trailer_match.group(1)), _norm_uuid(trailer_match.group(2))
            if trailer_orig and trailer_curr and trailer_curr != trailer_orig:
                indicators['TrailerIDChange'] = {'from': trailer_orig, 'to': trailer_curr}
        
        # --- Date Mismatch ---
        info_dates = dict(re.findall(r"/(ModDate|CreationDate)\s*\(\s*D:(\d{8,14})", txt))
        xmp_dates = {k: v for k, v in re.findall(r"<xmp:(ModifyDate|CreateDate)>([^<]+)</xmp:\1>", txt)}

        def _short(d: str) -> str: return re.sub(r"[-:TZ]", "", d)[:14]

        if "CreationDate" in info_dates and "CreateDate" in xmp_dates and _short(info_dates["CreationDate"]) != _short(xmp_dates["CreateDate"]):
            indicators['CreateDateMismatch'] = {'info': info_dates["CreationDate"], 'xmp': xmp_dates["CreateDate"]}
        if "ModDate" in info_dates and "ModifyDate" in xmp_dates and _short(info_dates["ModDate"]) != _short(xmp_dates["ModifyDate"]):
            indicators['ModifyDateMismatch'] = {'info': info_dates["ModDate"], 'xmp': xmp_dates["ModifyDate"]}
        
        return indicators

    def get_flag(self, indicators_dict, is_revision, parent_id=None):
        """
        Determines the file's status flag based on the found indicator keys.
        """
        if is_revision:
            return self._("revision_of").format(id=parent_id)

        keys_set = set(indicators_dict.keys())
        YES = "YES" if self.language.get() == "en" else "JA"
        NO = self._("status_no")

        high_risk_indicators = {
            "HasRevisions",
            "TouchUp_TextEdit",
            "Signature: Invalid",
        }

        if any(ind in high_risk_indicators for ind in keys_set):
            return YES

        if indicators_dict:
            return "Possible" if self.language.get() == "en" else "Sandsynligt"

        return NO
        
    
    
    def show_about(self):
        """Displays the 'About' popup window."""
        about_popup = Toplevel(self.root)
        about_popup.title(self._("menu_about"))
        about_popup.geometry("520x480") # Adjusted height slightly
        about_popup.resizable(True, True)
        about_popup.transient(self.root)

        # --- Layout ---
        outer_frame = ttk.Frame(about_popup, padding=10)
        outer_frame.pack(fill="both", expand=True)
        outer_frame.rowconfigure(0, weight=1)
        outer_frame.columnconfigure(0, weight=1)

        text_frame = ttk.Frame(outer_frame)
        text_frame.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side="right", fill="y")
        
        about_text_widget = tk.Text(text_frame, wrap="word", yscrollcommand=scrollbar.set, borderwidth=0, highlightthickness=0, background=about_popup.cget("background"))
        about_text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=about_text_widget.yview)

        # --- Text Formatting Tags ---
        about_text_widget.tag_configure("bold", font=("Segoe UI", 9, "bold"))
        about_text_widget.tag_configure("link", foreground="blue", underline=True)
        about_text_widget.tag_configure("header", font=("Segoe UI", 9, "bold", "underline"))

        # --- Make links clickable ---
        def _open_link(event):
            index = about_text_widget.index(f"@{event.x},{event.y}")
            tag_indices = about_text_widget.tag_ranges("link")
            for start, end in zip(tag_indices[0::2], tag_indices[1::2]):
                if about_text_widget.compare(start, "<=", index) and about_text_widget.compare(index, "<", end):
                    url = about_text_widget.get(start, end).strip()
                    if not url.startswith("http"):
                        url = "https://" + url
                    webbrowser.open(url)
                    break

        about_text_widget.tag_bind("link", "<Enter>", lambda e: about_text_widget.config(cursor="hand2"))
        about_text_widget.tag_bind("link", "<Leave>", lambda e: about_text_widget.config(cursor=""))
        about_text_widget.tag_bind("link", "<Button-1>", _open_link)

        # --- Insert Content ---
        about_text_widget.insert("end", f"{self._('about_version')} ({datetime.now().strftime('%d-%m-%Y')})\n", "bold")
        about_text_widget.insert("end", self._("about_developer_info"))

        # Add project website
        about_text_widget.insert("end", self._("about_project_website"), "bold")
        about_text_widget.insert("end", "github.com/Rasmus-Riis/PDFRecon\n", "link")

        about_text_widget.insert("end", "\n------------------------------------\n\n")
        
        about_text_widget.insert("end", self._("about_purpose_header") + "\n", "header")
        about_text_widget.insert("end", self._("about_purpose_text"))
        
        about_text_widget.insert("end", self._("about_included_software_header") + "\n", "header")
        about_text_widget.insert("end", self._("about_included_software_text").format(tool="ExifTool"))
        
        about_text_widget.insert("end", self._("about_website").format(tool="ExifTool"), "bold")
        about_text_widget.insert("end", "exiftool.org\n", "link")
        
        about_text_widget.insert("end", self._("about_source").format(tool="ExifTool"), "bold")
        about_text_widget.insert("end", "github.com/exiftool/exiftool\n", "link")
        
        about_text_widget.config(state="disabled") # Make read-only
        
        # --- Close Button ---
        close_button = ttk.Button(outer_frame, text=self._("close_button_text"), command=about_popup.destroy)
        close_button.grid(row=1, column=0, pady=(10, 0))
    
    def _check_for_updates(self):
        """Checks for a new version of the application on GitHub."""
        # Run the network request in a separate thread to not freeze the GUI
        threading.Thread(target=self._perform_update_check, daemon=True).start()

    def _perform_update_check(self):
        """The actual logic for the update check, meant to be run in a thread."""
        # --- IMPORTANT ---
        # Change this to your GitHub repository in the format "owner/repo"
        GITHUB_REPO = "Rasmus-Riis/PDFRecon"
        
        api_url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
        
        try:
            response = requests.get(api_url, timeout=10)
            response.raise_for_status()  # Raise an exception for bad status codes (4xx or 5xx)
            
            latest_release = response.json()
            latest_version_str = latest_release.get("tag_name", "").lstrip('v')
            
            if not latest_version_str:
                self.root.after(0, lambda: messagebox.showwarning(self._("update_error_title"), self._("update_parse_error_msg")))
                return

            # Compare versions. Converts "16.9.0" to (16, 9, 0) for proper comparison.
            current_version_tuple = tuple(map(int, self.app_version.split('.')))
            latest_version_tuple = tuple(map(int, latest_version_str.split('.')))

            if latest_version_tuple > current_version_tuple:
                release_url = latest_release.get("html_url")
                message = self._("update_available_msg").format(
                    new_version=latest_version_str,
                    current_version=self.app_version
                )
                if messagebox.askyesno(self._("update_available_title"), message):
                    webbrowser.open(release_url)
            else:
                self.root.after(0, lambda: messagebox.showinfo(self._("update_no_new_title"), self._("update_no_new_msg")))

        except requests.exceptions.RequestException as e:
            logging.error(f"Update check failed: {e}")
            self.root.after(0, lambda: messagebox.showerror(self._("update_error_title"), self._("update_net_error_msg")))
    
    def show_manual(self):
        """Displays a pop-up window with the program manual."""
        manual_popup = Toplevel(self.root)
        manual_popup.title(self._("manual_title"))
        
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        w, h = int(sw * 0.7), int(sh * 0.8)
        x, y = (sw - w) // 2, (sh - h) // 2
        manual_popup.geometry(f"{w}x{h}+{x}+{y}")
        
        manual_popup.resizable(True, True)
        manual_popup.transient(self.root)

        outer_frame = ttk.Frame(manual_popup, padding=10)
        outer_frame.pack(fill="both", expand=True)
        outer_frame.rowconfigure(0, weight=1)
        outer_frame.columnconfigure(0, weight=1)

        text_frame = ttk.Frame(outer_frame)
        text_frame.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side="right", fill="y")
        
        manual_text_widget = tk.Text(text_frame, wrap="word", yscrollcommand=scrollbar.set, borderwidth=0, highlightthickness=0, background=manual_popup.cget("background"), font=("Segoe UI", 10))
        manual_text_widget.pack(side="left", fill="both", expand=True, padx=5)
        scrollbar.config(command=manual_text_widget.yview)

        manual_text_widget.tag_configure("h1", font=("Segoe UI", 16, "bold", "underline"), spacing3=10)
        manual_text_widget.tag_configure("h2", font=("Segoe UI", 12, "bold"), spacing1=10, spacing3=5)
        manual_text_widget.tag_configure("b", font=("Segoe UI", 10, "bold"))
        manual_text_widget.tag_configure("i", font=("Segoe UI", 10, "italic"))
        manual_text_widget.tag_configure("red", foreground="#C00000")
        manual_text_widget.tag_configure("yellow", foreground="#C07000")
        manual_text_widget.tag_configure("green", foreground="#008000")

        full_manual_text = self._("full_manual")
        
        for line in full_manual_text.strip().split('\n'):
            line = line.strip()
            if line.startswith("# "):
                manual_text_widget.insert(tk.END, line[2:] + "\n", "h1")
            elif line.startswith("## "):
                manual_text_widget.insert(tk.END, line[3:] + "\n", "h2")
            else:
                parts = re.split(r'(<.*?>)', line)
                active_tags = set()
                for part in parts:
                    if part.startswith("</"):
                        tag_name = part[2:-1]
                        if tag_name in active_tags:
                            active_tags.remove(tag_name)
                    elif part.startswith("<"):
                        tag_name = part[1:-1]
                        active_tags.add(tag_name)
                    elif part:
                        manual_text_widget.insert(tk.END, part, tuple(active_tags))
                manual_text_widget.insert(tk.END, "\n")

        manual_text_widget.config(state="disabled")

        close_button = ttk.Button(outer_frame, text=self._("close_button_text"), command=manual_popup.destroy)
        close_button.grid(row=1, column=0, pady=(10, 0))
        
    
    def _process_and_validate_revisions(self, potential_revisions, original_fp, scan_root_folder):
        """
        Processes, validates, and compares a list of potential PDF revisions.
        Returns a list of dictionaries for valid revisions that should be reported.
        """
        valid_revision_results = []
        invalid_xref_dir = scan_root_folder / "Invalid XREF"

        for rev_path, basefile, rev_raw in potential_revisions:
            tmp_path = None
            try:
                # Use a temporary file to run ExifTool without writing to the final destination first
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
                    tmp_file.write(rev_raw)
                    tmp_path = Path(tmp_file.name)
                
                rev_exif = self.exiftool_output(tmp_path, detailed=True)

                # Conditionally copy revisions with invalid XREF tables if the setting is enabled
                if PDFReconConfig.EXPORT_INVALID_XREF and "Warning" in rev_exif and "Invalid xref table" in rev_exif:
                    logging.info(f"Submitting invalid XREF revision for {basefile} to be copied.")
                    dest_path = invalid_xref_dir / rev_path.name
                    if self.copy_executor:
                        self.copy_executor.submit(self._perform_copy, rev_raw, dest_path)
                    continue  # Skip adding this invalid revision to the results

                # If valid, write the revision to its final destination
                rev_path.parent.mkdir(exist_ok=True)
                rev_path.write_bytes(rev_raw)
                
                rev_md5 = hashlib.md5(rev_raw).hexdigest()
                rev_txt = self.extract_text(rev_raw)
                revision_timeline = self.generate_comprehensive_timeline(rev_path, rev_txt, rev_exif)

                # Perform a visual comparison to see if the revision is identical to the original
                is_identical = False
                try:
                    with fitz.open(original_fp) as doc_orig, fitz.open(rev_path) as doc_rev:
                        pages_to_compare = min(doc_orig.page_count, doc_rev.page_count, PDFReconConfig.VISUAL_DIFF_PAGE_LIMIT)
                        if pages_to_compare > 0:
                            is_identical = True
                            for i in range(pages_to_compare):
                                page_orig, page_rev = doc_orig.load_page(i), doc_rev.load_page(i)
                                if page_orig.rect != page_rev.rect: is_identical = False; break
                                pix_orig, pix_rev = page_orig.get_pixmap(dpi=96), page_rev.get_pixmap(dpi=96)
                                img_orig = Image.frombytes("RGB", [pix_orig.width, pix_orig.height], pix_orig.samples)
                                img_rev = Image.frombytes("RGB", [pix_rev.width, pix_rev.height], pix_rev.samples)
                                if ImageChops.difference(img_orig, img_rev).getbbox() is not None: is_identical = False; break
                except Exception as e:
                    logging.warning(f"Could not visually compare {rev_path.name}, keeping it. Error: {e}")
                
                revision_row_data = { 
                    "path": rev_path, "indicator_keys": {"Revision": {}}, "md5": rev_md5, "exif": rev_exif, 
                    "is_revision": True, "timeline": revision_timeline, "original_path": original_fp, 
                    "is_identical": is_identical, "status": "success"
                }
                valid_revision_results.append(revision_row_data)
            finally:
                if tmp_path and tmp_path.exists():
                    os.remove(tmp_path)
        
        return valid_revision_results

    def _process_single_file(self, fp, scan_root_folder):
        """
        Processes a single PDF file, submitting copy jobs to a dedicated thread pool.
        """
        try:
            self.validate_pdf_file(fp)

            raw = fp.read_bytes()
            doc = fitz.open(stream=raw, filetype="pdf")
            txt = self.extract_text(raw)
            indicators = self.detect_indicators(fp, txt, doc)
            md5_hash = md5_file(fp)
            exif = self.exiftool_output(fp, detailed=True)

            # Conditionally submit invalid XREF originals to the copy pool if setting is enabled
            if PDFReconConfig.EXPORT_INVALID_XREF and "Warning" in exif and "Invalid xref table" in exif:
                invalid_xref_dir = scan_root_folder / "Invalid XREF"
                dest_path = invalid_xref_dir / fp.name
                if self.copy_executor:
                    self.copy_executor.submit(self._perform_copy, fp, dest_path)

            tool_change_info = self._detect_tool_change_from_exif(exif)
            if tool_change_info.get("changed"):
                indicators['ToolChange'] = {}
            original_timeline = self.generate_comprehensive_timeline(fp, txt, exif)
            potential_revisions = self.extract_revisions(raw, fp)
            doc.close()
            
            self._add_layer_indicators(raw, fp, indicators)

            # Delegate all revision processing to the new helper method
            valid_revision_results = self._process_and_validate_revisions(potential_revisions, fp, scan_root_folder)

            if valid_revision_results:
                indicators['HasRevisions'] = {'count': len(valid_revision_results)}

            original_row_data = {
                "path": fp, "indicator_keys": indicators, "md5": md5_hash, 
                "exif": exif, "is_revision": False, "timeline": original_timeline, "status": "success"
            }
            
            results = [original_row_data] + valid_revision_results
            return results

        except PDFTooLargeError as e:
            logging.warning(f"Skipping large file {fp.name}: {e}")
            return [{"path": fp, "status": "error", "error_type": "file_too_large", "error_message": str(e)}]
        except PDFEncryptedError as e:
            logging.warning(f"Skipping encrypted file {fp.name}: {e}")
            return [{"path": fp, "status": "error", "error_type": "file_encrypted", "error_message": str(e)}]
        except PDFCorruptionError as e:
            logging.warning(f"Skipping corrupt file {fp.name}: {e}")
            return [{"path": fp, "status": "error", "error_type": "file_corrupt", "error_message": str(e)}]
        except Exception as e:
            logging.exception(f"Unexpected error processing file {fp.name}")
            return [{"path": fp, "status": "error", "error_type": "processing_error", "error_message": str(e)}]        
    def _prompt_and_export(self, file_format):
        """Prompts the user for a file path and calls the relevant export function."""
        if not self.report_data:
            messagebox.showwarning(self._("no_data_to_save_title"), self._("no_data_to_save_message"))
            return
        
        file_types = {
            "xlsx": [("Excel files", "*.xlsx")], "csv": [("CSV files", "*.csv")],
            "json": [("JSON files", "*.json")], "html": [("HTML files", "*.html")]
        }
        file_path = filedialog.asksaveasfilename(defaultextension=f".{file_format}", filetypes=file_types[file_format])
        if not file_path: return

        try:
            export_methods = {
                "xlsx": self._export_to_excel, "csv": self._export_to_csv,
                "json": self._export_to_json, "html": self._export_to_html
            }
            export_methods[file_format](file_path)
            
            # Ask to open the containing folder
            if messagebox.askyesno(self._("excel_saved_title"), self._("excel_saved_message")):
                webbrowser.open(os.path.dirname(file_path))

        except Exception as e:
            logging.error(f"Error exporting to {file_format.upper()}: {e}")
            messagebox.showerror(self._("excel_save_error_title"), self._("excel_save_error_message").format(e=e))

    def _export_to_excel(self, file_path):
        """Exports the displayed data to XLSX with a frozen header and word wrap enabled."""
        import logging
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        logging.info(f"Exporting report to Excel file: {file_path}")

        wb = Workbook()
        ws = wb.active
        ws.title = "PDFRecon Results"

        headers = [self._(key) for key in self.columns_keys]
        if len(headers) >= 10:
            headers[9] = f"{self._('col_indicators')} {self._('excel_indicators_overview')}"

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=self._clean_cell_value(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        
        ws.freeze_panes = 'A2'

        # --- OPTIMIZATION ---
        # Create a lookup dictionary once to avoid repeated searches in the main loop.
        # This significantly improves performance for large datasets.
        indicators_by_path = {}
        for item in getattr(self, "all_scan_data", []):
            path_str = str(item.get("path"))
            indicator_dict = item.get("indicator_keys") or {}
            if indicator_dict:
                lines = [self._format_indicator_details(key, details) for key, details in indicator_dict.items()]
                indicators_by_path[path_str] = "• " + "\n• ".join(lines)
            else:
                indicators_by_path[path_str] = ""

        for row_idx, row_data in enumerate(getattr(self, "report_data", []), start=2):
            try:
                path = row_data[4] # Path is now at index 4
            except IndexError:
                path = ""

            exif_text = self.exif_outputs.get(path, "")
            # Use the fast lookup dictionary instead of the slow nested function
            indicators_full = indicators_by_path.get(path, "")
            note_text = self.file_annotations.get(path, "")

            row_out = list(row_data)
            
            while len(row_out) < len(headers):
                row_out.append("")
            
            row_out[8] = exif_text         # EXIF is at index 8
            if indicators_full:
                row_out[9] = indicators_full # Indicators is at index 9
            row_out[10] = note_text        # Note is at index 10

            for col_idx, value in enumerate(row_out, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=self._clean_cell_value(value))
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col in ws.columns:
            try:
                max_len = max(len(str(c.value).split('\n')[0]) for c in col if c.value)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
            except (ValueError, TypeError):
                pass

        wb.save(file_path)
        
    def _export_to_csv(self, file_path):
        """Exports the displayed data to a CSV file."""
        headers = [self._(key) for key in self.columns_keys]
        
        def _indicators_for_path(path_str: str) -> str:
            """Helper function to get a semicolon-separated string of indicators."""
            rec = next((d for d in self.all_scan_data if str(d.get('path')) == path_str), None)
            if not rec: return ""
            indicator_dict = rec.get('indicator_keys') or {}
            if not indicator_dict: return ""

            lines = [self._format_indicator_details(key, details) for key, details in indicator_dict.items()]
            return "; ".join(lines)

        # Prepare data with full EXIF output + full indicators
        data_for_export = []
        for row_data in self.report_data:
            new_row = row_data[:]
            path = new_row[3]
            exif_output = self.exif_outputs.get(path, "")
            new_row[7] = exif_output  # Replace "Click to view." with actual output
            indicators_full = _indicators_for_path(path)
            if indicators_full:
                new_row[8] = indicators_full
            note_text = self.file_annotations.get(path, "")
            while len(new_row) < len(headers):
                new_row.append("")
            new_row[9] = note_text
  
            data_for_export.append(new_row)

        # Use utf-8-sig for better Excel compatibility with special characters
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data_for_export)

    def _export_to_json(self, file_path):
        """Exports a more detailed report of all scanned data and notes to a JSON file."""
        scan_data_export = []
        for item in self.all_scan_data:
            path_str = str(item['path'])
            item_copy = item.copy()
            item_copy['path'] = path_str # Convert Path object to string
            if 'original_path' in item_copy:
                item_copy['original_path'] = str(item_copy['original_path'])
            
            if 'indicator_keys' in item_copy:
                serializable_indicators = {}
                for key, details in item_copy['indicator_keys'].items():
                    if 'fonts' in details:
                        serializable_details = details.copy()
                        serializable_details['fonts'] = {k: list(v) for k, v in details['fonts'].items()}
                        serializable_indicators[key] = serializable_details
                    else:
                        serializable_indicators[key] = details
                item_copy['indicator_keys'] = serializable_indicators

            item_copy['exif_data'] = self.exif_outputs.get(path_str, "")
            scan_data_export.append(item_copy)
        
        full_export_payload = {
            'scan_results': scan_data_export,
            'file_annotations': self.file_annotations
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(full_export_payload, f, indent=4, default=str)
            
    def _export_to_html(self, file_path):
        """Exports a simple, color-coded HTML report."""
        import html
        html_template = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <title>PDFRecon Report</title>
            <style>
                body {{ font-family: sans-serif; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; word-break: break-all; }}
                th {{ background-color: #f2f2f2; }}
                .red-row {{ background-color: #FFDDDD; }}
                .yellow-row {{ background-color: #FFFFCC; }}
                .blue-row {{ background-color: #CCE5FF; }}
                .gray-row {{ background-color: #E0E0E0; }}
            </style>
        </head>
        <body>
            <h1>PDFRecon Report</h1>
            <p>Generated on {date}</p>
            <table>
                <thead><tr>{headers}</tr></thead>
                <tbody>{rows}</tbody>
            </table>
        </body>
        </html>
        """
        headers = "".join(f"<th>{self._(key)}</th>" for key in self.columns_keys)
        rows = ""
        tag_map = {"red_row": "red-row", "yellow_row": "yellow-row", "blue_row": "blue-row", "gray_row": "gray-row"}
        
        # --- Generate Table Rows ---
        for i, values in enumerate(self.report_data):
            tag_class = ""
            try:
                matching_id = next((item_id for item_id in self.tree.get_children() if self.tree.item(item_id, "values")[3] == values[3]), None)
                if matching_id:
                    tags = self.tree.item(matching_id, "tags")
                    if tags:
                        tag_class = tag_map.get(tags[0], "")
            except (IndexError, StopIteration):
                 pass
            
            path_str = values[3]
            note_text = html.escape(self.file_annotations.get(path_str, "")).replace('\n', '<br>')
            
            row_values = [html.escape(str(v)) for v in values]
            while len(row_values) < len(self.columns_keys):
                row_values.append("")
            row_values[9] = note_text

            rows += f'<tr class="{tag_class}">' + "".join(f"<td>{v}</td>" for v in row_values) + "</tr>"

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_template.format(
                date=datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
                headers=headers,
                rows=rows
            ))
   
    def _extract_touchup_text(self, doc):
        """
        Parses a PDF's internal objects to find any text associated with a TouchUp_TextEdit flag.
        Returns a list of extracted text strings.
        """
        found_text = []
        if not doc or doc.is_closed:
            return found_text
            
        # Iterate through all objects in the PDF's cross-reference table
        for xref in range(1, doc.xref_length()):
            try:
                # Get the raw source code of the object
                obj_source = doc.xref_object(xref, compressed=False)
                
                # Check if this object contains the flag
                if "/TouchUp_TextEdit" in obj_source:
                    # If it's a stream object, get its decoded content
                    stream = doc.xref_stream(xref)
                    if stream:
                        # This regex finds text within parentheses followed by the 'Tj' (Show Text) operator.
                        # It's a common pattern for simple text entries in content streams.
                        matches = re.findall(rb"\((.*?)\)\s*Tj", stream)
                        for match in matches:
                            try:
                                # Decode the text using a PDF-specific utility to handle special characters
                                decoded_text = fitz.utils.pdfdoc_decode(match)
                                if decoded_text.strip():
                                    found_text.append(decoded_text.strip())
                            except:
                                continue # Ignore text that can't be decoded
            except Exception as e:
                logging.warning(f"Could not process object {xref} for TouchUp text: {e}")
                continue
        return found_text
    
    def _get_text_for_comparison(self, source):
        """
        Performs a robust, layout-preserving text extraction on a PDF.
        Source can be bytes, a string path, or a Path object.
        """
        full_text = []
        doc = None
        try:
            if isinstance(source, bytes):
                doc = fitz.open(stream=source, filetype="pdf")
            else:
                resolved_path = self._resolve_case_path(source) # Resolve the path
                doc = fitz.open(resolved_path)

            for page in doc:
                full_text.append(page.get_text("text", sort=True))
            return "\n".join(full_text)
        except Exception as e:
            logging.error(f"Robust text extraction failed: {e}")
            return ""
        finally:
            if doc:
                doc.close()    
        
            
    def show_pdf_viewer_popup(self, item_id):
        """Displays a simple PDF viewer for the selected file."""
        self.root.config(cursor="watch")
        self.root.update_idletasks()

        try:
            path_str = self.tree.item(item_id, "values")[4]
            file_name = self.tree.item(item_id, "values")[1]
            resolved_path = self._resolve_case_path(path_str)
        except (IndexError, TypeError):
            self.root.config(cursor="")
            return

        try:
            # --- Popup Window Setup ---
            popup = Toplevel(self.root)
            popup.title(f"{self._('pdf_viewer_title')} - {file_name}")
            
            # --- PDF Loading ---
            popup.current_page = 0
            popup.doc = fitz.open(resolved_path) # Use resolved_path
            popup.total_pages = len(popup.doc)
            # --- Widget Layout ---
            main_frame = ttk.Frame(popup, padding=10)
            main_frame.pack(fill="both", expand=True)
            main_frame.rowconfigure(0, weight=1)
            main_frame.columnconfigure(0, weight=1)

            image_label = ttk.Label(main_frame)
            image_label.grid(row=0, column=0, pady=5, sticky="nsew")
            
            nav_frame = ttk.Frame(main_frame)
            nav_frame.grid(row=1, column=0, pady=(10,0))
            
            prev_button = ttk.Button(nav_frame, text=self._("diff_prev_page"))
            page_label = ttk.Label(nav_frame, text="", font=("Segoe UI", 9, "italic"))
            next_button = ttk.Button(nav_frame, text=self._("diff_next_page"))

            prev_button.pack(side="left", padx=10)
            page_label.pack(side="left", padx=10)
            next_button.pack(side="left", padx=10)
            
            def update_page(page_num):
                """Renders and displays a specific page of the PDF."""
                if not (0 <= page_num < popup.total_pages): return
                
                popup.current_page = page_num
                self.root.config(cursor="watch")
                self.root.update()

                # Render page to a pixmap, then to a PIL Image
                page = popup.doc.load_page(page_num)
                pix = page.get_pixmap(dpi=150)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                
                # Scale image to fit the window
                max_img_w, max_img_h = main_frame.winfo_width() * 0.95, main_frame.winfo_height() * 0.9
                img_w, img_h = img.size
                ratio = min(max_img_w / img_w, max_img_h / img_h) if img_w > 0 and img_h > 0 else 1
                scaled_size = (int(img_w * ratio), int(img_h * ratio))

                # Convert to Tkinter PhotoImage and display
                img_tk = ImageTk.PhotoImage(img.resize(scaled_size, Image.Resampling.LANCZOS))
                popup.img_tk = img_tk # Keep a reference to prevent garbage collection
                
                image_label.config(image=img_tk)
                page_label.config(text=self._("diff_page_label").format(current=page_num + 1, total=popup.total_pages))
                prev_button.config(state="normal" if page_num > 0 else "disabled")
                next_button.config(state="normal" if page_num < popup.total_pages - 1 else "disabled")
                self.root.config(cursor="")

            prev_button.config(command=lambda: update_page(popup.current_page - 1))
            next_button.config(command=lambda: update_page(popup.current_page + 1))
            
            def on_close():
                """Ensures the PDF document is closed properly."""
                if hasattr(popup, 'doc') and popup.doc:
                    popup.doc.close()
                popup.destroy()
            popup.protocol("WM_DELETE_WINDOW", on_close)

            sw = self.root.winfo_screenwidth()
            sh = self.root.winfo_screenheight()
            w, h = int(sw * 0.7), int(sh * 0.85)
            x, y = (sw - w) // 2, (sh - h) // 2
            popup.geometry(f"{w}x{h}+{x}+{y}")

            update_page(0)
            
            popup.transient(self.root)
            popup.grab_set()

        except Exception as e:
            logging.error(f"PDF viewer error: {e}")
            messagebox.showerror(self._("pdf_viewer_error_title"), self._("pdf_viewer_error_message").format(e=e), parent=self.root)
            if 'popup' in locals() and hasattr(popup, 'doc') and popup.doc:
                popup.doc.close()
        finally:
            self.root.config(cursor="")
            
    
    def _format_indicator_details(self, key, details):
        """Generates a human-readable string for an indicator and its details."""
        if key == 'TouchUp_TextEdit':
            found_text_str = ""
            diff_str = ""
            if details and details.get('found_text'):
                text_list_str = ', '.join(f'"{t}"' for t in details['found_text'])
                found_text_str = f"Found isolated text: {text_list_str}"
            if details and details.get('text_diff'):
                diff_str = "A text comparison to a previous version is available."
            
            if found_text_str and diff_str:
                return f"TouchUp TextEdit ({found_text_str}. {diff_str})"
            elif found_text_str:
                return f"TouchUp TextEdit ({found_text_str})"
            elif diff_str:
                return f"TouchUp TextEdit ({diff_str})"
            else:
                return "TouchUp TextEdit (Flag found, but no specific text or revisions to compare)"

        if key == 'MultipleCreators':
            return f"Multiple Creators (Found {details['count']}): " + ", ".join(f'"{v}"' for v in details['values'])
        if key == 'MultipleProducers':
            return f"Multiple Producers (Found {details['count']}): " + ", ".join(f'"{v}"' for v in details['values'])
        if key == 'MultipleFontSubsets':
            font_details = []
            for base_font, subsets in details['fonts'].items():
                font_details.append(f"'{base_font}': {{{{{', '.join(subsets)}}}}}")
            return f"Multiple Font Subsets: " + "; ".join(font_details)
        if key == 'CreateDateMismatch':
            return f"Creation Date Mismatch: Info='{details['info']}', XMP='{details['xmp']}'"
        if key == 'ModifyDateMismatch':
            return f"Modify Date Mismatch: Info='{details['info']}', XMP='{details['xmp']}'"
        if key == 'TrailerIDChange':
            return f"Trailer ID Changed: From [{details['from']}] to [{details['to']}]"
        if key == 'XMPIDChange':
            return f"XMP DocumentID Changed: From [{details['from']}] to [{details['to']}]"
        if key == 'MultipleStartxref':
            return f"Multiple startxref (Found {details['count']})"
        if key == 'IncrementalUpdates':
            return f"Incremental updates (Found {details['count']} versions)"
        if key == 'ObjGenGtZero':
            return f"Objects with generation > 0 (Found {details['count']} objects)"
        if key == 'HasLayers':
            return f"Has Layers (Found {details['count']})"
        if key == 'MoreLayersThanPages':
            return f"More Layers ({details['layers']}) Than Pages ({details['pages']})"
        # Fallback for simple indicators with no details
        return key.replace("_", " ")
        
if __name__ == "__main__":
    # Ensure multiprocessing works correctly when the app is frozen (e.g., by PyInstaller).
    if sys.platform.startswith('win'):
        multiprocessing.freeze_support()
        
    # Initialize the main Tkinter window with DnD support
    root = TkinterDnD.Tk()
    app = PDFReconApp(root)
    # Start the application's main event loop
    root.mainloop()

